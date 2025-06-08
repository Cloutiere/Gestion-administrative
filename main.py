import datetime
import os
from functools import wraps
from typing import Any  # Importation de Any pour les types génériques

import openpyxl
import psycopg2
import psycopg2.extras
from flask import Flask, flash, g, jsonify, redirect, render_template, request, url_for
from flask_login import (
    LoginManager,
    current_user,
    login_required,
    login_user,
    logout_user,
)
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

# Importation du modèle User et des fonctions de hachage de mot de passe
from models import User, check_hashed_password, hash_password

# --- Configuration de l'application Flask ---
app = Flask(__name__)
app.secret_key = os.urandom(24)  # Clé secrète pour la session Flask
UPLOAD_FOLDER = "uploads"
ALLOWED_EXTENSIONS = {"xlsx"}
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# --- Configuration de Flask-Login ---
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"  # La vue vers laquelle rediriger si l'utilisateur n'est pas connecté
login_manager.login_message = "Veuillez vous connecter pour accéder à cette page."
login_manager.login_message_category = "info"


@login_manager.user_loader
def load_user(user_id: int) -> User | None:
    """
    Fonction de rappel utilisée par Flask-Login pour recharger l'objet utilisateur
    à partir de l'ID stocké dans la session.
    """
    return get_user_by_id(user_id)


# --- Filtre Jinja2 personnalisé pour formater les périodes ---
def format_periodes_filter(value: float | None) -> str:
    """
    Formate un nombre de périodes pour l'affichage dans les templates Jinja2.
    Affiche les décimales uniquement si elles sont non nulles, et évite les zéros superflus.
    Exemples: 5.00 -> "5", 1.50 -> "1.5", 0.75 -> "0.75".
    """
    if value is None:
        return ""
    # Convertir en float pour s'assurer que les opérations de formatage sont cohérentes.
    # Gère également les types Decimal potentiellement retournés par psycopg2.
    num = float(value)

    # Si le nombre est un entier (e.g., 5.0, 7.0), le convertir en chaîne d'entier.
    if num == int(num):
        return str(int(num))
    # Sinon, formater avec deux décimales, puis supprimer les zéros superflus à la fin.
    # .rstrip('0') enlève les zéros de fin, .rstrip('.') enlève le point si tous les zéros ont été enlevés.
    formatted_str = f"{num:.2f}"
    return formatted_str.rstrip("0").rstrip(".")


app.jinja_env.filters["format_periodes"] = format_periodes_filter


# --- Context Processor pour rendre current_user disponible dans tous les templates ---
@app.context_processor
def inject_global_data() -> dict[str, Any]:
    """
    Rend certaines variables globales disponibles dans tous les templates Jinja2.
    """
    return {
        "current_user": current_user,
        "SCRIPT_YEAR": datetime.datetime.now().year,
    }


# --- Configuration de la base de données ---
DB_HOST = os.environ.get("PGHOST")
DB_NAME = os.environ.get("PGDATABASE")
DB_USER = os.environ.get("PGUSER")
DB_PASS = os.environ.get("PGPASSWORD")
DB_PORT = os.environ.get("PGPORT", "5432")


def get_db_connection_string() -> str:
    """Construit la chaîne de connexion à la base de données en utilisant les variables d'environnement."""
    return f"dbname='{DB_NAME}' user='{DB_USER}' host='{DB_HOST}' password='{DB_PASS}' port='{DB_PORT}'"


def get_db():
    """
    Ouvre une nouvelle connexion à la base de données si ce n'est pas déjà fait pour la requête actuelle.
    La connexion est stockée dans l'objet 'g' de Flask pour être réutilisée durant le cycle de requête.
    """
    if "db" not in g:
        try:
            conn_string = get_db_connection_string()
            g.db = psycopg2.connect(conn_string)
        except psycopg2.Error as e:
            app.logger.error(f"Erreur de connexion à la base de données: {e}")
            g.db = None
    return g.db


@app.teardown_appcontext
def close_db(_exception=None):
    """
    Ferme la connexion à la base de données à la fin de la requête Flask.
    """
    db = g.pop("db", None)
    if db is not None and not db.closed:
        try:
            db.close()
        except psycopg2.Error as e:
            app.logger.error(f"Erreur lors de la fermeture de la connexion DB: {e}")


# --- Fonctions d'accès aux données (DAO - Data Access Object) pour les utilisateurs ---


def get_user_by_id(user_id: int) -> User | None:
    """
    Récupère un utilisateur par son ID depuis la base de données,
    y compris ses permissions d'accès aux champs.
    """
    db = get_db()
    if not db:
        return None
    try:
        with db.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            cur.execute(
                """
                SELECT u.id, u.username, u.is_admin, uca.champ_no
                FROM Users u
                LEFT JOIN user_champ_access uca ON u.id = uca.user_id
                WHERE u.id = %s;
            """,
                (user_id,),
            )
            user_data_rows = cur.fetchall()

            if not user_data_rows:
                return None

            first_row = user_data_rows[0]
            _id = first_row["id"]
            username = first_row["username"]
            is_admin = first_row["is_admin"]

            allowed_champs: list[str] = []
            if not is_admin:  # Si l'utilisateur n'est PAS admin, collecter ses champs autorisés
                for row in user_data_rows:
                    if row["champ_no"]:  # S'assurer qu'il y a un champ et que la jointure n'est pas nulle
                        allowed_champs.append(row["champ_no"])
                allowed_champs = list(set(allowed_champs))  # Éliminer les doublons
            # Si is_admin est True, allowed_champs reste vide, ce qui est géré par User.can_access_champ.

            return User(_id=_id, username=username, is_admin=is_admin, allowed_champs=allowed_champs)
    except psycopg2.Error as e:
        app.logger.error(f"Erreur DAO get_user_by_id pour {user_id}: {e}")
        return None


def get_user_by_username(username: str) -> dict | None:
    """
    Récupère un utilisateur par son nom d'utilisateur depuis la base de données.
    Retourne un dictionnaire incluant le hash du mot de passe.
    """
    db = get_db()
    if not db:
        return None
    try:
        with db.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            cur.execute("SELECT id, username, password_hash, is_admin FROM Users WHERE username = %s;", (username,))
            user_data = cur.fetchone()
            return dict(user_data) if user_data else None
    except psycopg2.Error as e:
        app.logger.error(f"Erreur DAO get_user_by_username pour {username}: {e}")
        if db and not db.closed:
            db.rollback()
        return None


def get_users_count() -> int:
    """
    Compte le nombre total d'utilisateurs dans la base de données.
    """
    db = get_db()
    if not db:
        return 0
    try:
        with db.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM Users;")
            return cur.fetchone()[0]
    except psycopg2.Error as e:
        app.logger.error(f"Erreur DAO get_users_count: {e}")
        return 0


def create_user(username: str, password: str, is_admin: bool = False) -> User | None:
    """
    Crée un nouvel utilisateur dans la base de données avec un mot de passe haché.
    Retourne l'objet User créé ou None en cas d'échec (ex: nom d'utilisateur déjà pris).
    """
    db = get_db()
    if not db:
        return None
    try:
        hashed_pwd = hash_password(password)
        with db.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            cur.execute(
                "INSERT INTO Users (username, password_hash, is_admin) VALUES (%s, %s, %s) RETURNING id;",
                (username, hashed_pwd, is_admin),
            )
            user_id = cur.fetchone()["id"]
            db.commit()  # Commit pour la création de l'utilisateur
            # La liste allowed_champs est vide ici car la gestion se fait ailleurs (admin ou premier user)
            return User(_id=user_id, username=username, is_admin=is_admin, allowed_champs=[])
    except psycopg2.errors.UniqueViolation:
        db.rollback()
        app.logger.warning(f"Tentative de création d'un utilisateur existant: {username}")
        return None
    except psycopg2.Error as e:
        app.logger.error(f"Erreur DAO create_user pour {username}: {e}")
        if db and not db.closed:
            db.rollback()
        return None


def get_all_users_with_access_info() -> list[dict]:
    """
    Récupère tous les utilisateurs avec leurs informations de base et la liste
    des champs auxquels ils ont accès.
    """
    db = get_db()
    if not db:
        return []
    try:
        with db.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            cur.execute(
                """
                SELECT u.id, u.username, u.is_admin, ARRAY_AGG(uca.champ_no ORDER BY uca.champ_no) AS allowed_champs
                FROM Users u
                LEFT JOIN user_champ_access uca ON u.id = uca.user_id
                GROUP BY u.id, u.username, u.is_admin
                ORDER BY u.username;
            """
            )
            users_data = []
            for row in cur.fetchall():
                user_dict = dict(row)
                # ARRAY_AGG retourne un tableau Python, il peut contenir None si pas d'accès
                if user_dict["is_admin"]:
                    user_dict["allowed_champs"] = []  # Les admins ont tout par défaut.
                else:
                    user_dict["allowed_champs"] = (
                        [c for c in user_dict["allowed_champs"] if c is not None] if user_dict["allowed_champs"] else []
                    )
                users_data.append(user_dict)
            return users_data
    except psycopg2.Error as e:
        app.logger.error(f"Erreur DAO get_all_users_with_access_info: {e}")
        if db and not db.closed:
            db.rollback()
        return []


def update_user_champ_access(user_id: int, champ_nos: list[str]) -> bool:
    """
    Met à jour les permissions d'accès aux champs pour un utilisateur spécifique.
    Remplace tous les accès existants par la nouvelle liste fournie.
    """
    db = get_db()
    if not db:
        return False
    try:
        with db.cursor() as cur:
            # Supprimer tous les accès existants pour cet utilisateur
            cur.execute("DELETE FROM user_champ_access WHERE user_id = %s;", (user_id,))
            # Insérer les nouveaux accès
            if champ_nos:
                values = [(user_id, champ_no) for champ_no in champ_nos]
                psycopg2.extras.execute_values(
                    cur,
                    "INSERT INTO user_champ_access (user_id, champ_no) VALUES %s ON CONFLICT (user_id, champ_no) DO NOTHING;",
                    values,
                )
            db.commit()
            return True
    except psycopg2.Error as e:
        db.rollback()
        app.logger.error(f"Erreur DAO update_user_champ_access pour user {user_id}: {e}")
        return False


def delete_user_data(user_id: int) -> bool:
    """
    Supprime un utilisateur de la base de données.
    Cela supprimera aussi ses accès aux champs grâce à la contrainte ON DELETE CASCADE.
    """
    db = get_db()
    if not db:
        return False
    try:
        with db.cursor() as cur:
            cur.execute("DELETE FROM Users WHERE id = %s;", (user_id,))
            db.commit()
            return cur.rowcount > 0
    except psycopg2.Error as e:
        db.rollback()
        app.logger.error(f"Erreur DAO delete_user_data pour user {user_id}: {e}")
        return False


def grant_access_to_all_champs(user_id: int) -> bool:
    """
    Accorde à un utilisateur l'accès à tous les champs existants dans la base de données.
    Utilisé typiquement pour le premier utilisateur administrateur.
    """
    db = get_db()
    if not db:
        return False
    try:
        with db.cursor() as cur:
            cur.execute("SELECT ChampNo FROM Champs;")
            all_champs = [row[0] for row in cur.fetchall()]
            if all_champs:
                values = [(user_id, champ_no) for champ_no in all_champs]
                psycopg2.extras.execute_values(
                    cur,
                    "INSERT INTO user_champ_access (user_id, champ_no) VALUES %s ON CONFLICT (user_id, champ_no) DO NOTHING;",
                    values,
                )
            db.commit()
            return True
    except psycopg2.Error as e:
        db.rollback()
        app.logger.error(f"Erreur DAO grant_access_to_all_champs pour user {user_id}: {e}")
        return False


# --- Décorateur personnalisé pour l'accès administrateur ---
def admin_required(f):
    """
    Décorateur qui restreint l'accès à la route aux seuls utilisateurs administrateurs.
    Doit être utilisé APRÈS @login_required.
    """

    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            flash("Vous n'avez pas les permissions suffisantes pour accéder à cette page.", "error")
            return redirect(url_for("index"))
        return f(*args, **kwargs)

    return decorated_function


# --- Fonctions d'accès aux données (DAO - Data Access Object) existantes ---


def get_all_champs() -> list[dict]:
    """
    Récupère tous les champs depuis la base de données, y compris leur statut de verrouillage.
    Les champs sont triés par numéro de champ.
    Retourne une liste de dictionnaires, chaque dictionnaire représentant un champ.
    """
    db = get_db()
    if not db:
        return []
    try:
        with db.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            cur.execute("SELECT ChampNo, ChampNom, EstVerrouille FROM Champs ORDER BY ChampNo;")
            return [dict(row) for row in cur.fetchall()]
    except psycopg2.Error as e:
        app.logger.error(f"Erreur DAO get_all_champs: {e}")
        if db and not db.closed:
            db.rollback()
        return []


def get_champ_details(champ_no: str) -> dict | None:
    """
    Récupère les détails d'un champ spécifique par son numéro, y compris son statut de verrouillage.
    Retourne un dictionnaire représentant le champ, ou None si non trouvé.
    """
    db = get_db()
    if not db:
        return None
    try:
        with db.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            cur.execute("SELECT ChampNo, ChampNom, EstVerrouille FROM Champs WHERE ChampNo = %s;", (champ_no,))
            champ_row = cur.fetchone()
        return dict(champ_row) if champ_row else None
    except psycopg2.Error as e:
        app.logger.error(f"Erreur DAO get_champ_details pour {champ_no}: {e}")
        if db and not db.closed:
            db.rollback()
        return None


def get_enseignants_par_champ(champ_no: str) -> list[dict]:
    """
    Récupère les enseignants associés à un champ spécifique.
    Les enseignants sont triés par statut fictif (les fictifs après les réels),
    puis par nom de famille, puis par prénom.
    Retourne une liste de dictionnaires.
    """
    db = get_db()
    if not db:
        return []
    try:
        with db.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            cur.execute(
                """
                SELECT EnseignantID, NomComplet, Nom, Prenom, EstTempsPlein, EstFictif, PeutChoisirHorsChampPrincipal
                FROM Enseignants WHERE ChampNo = %s ORDER BY EstFictif, Nom, Prenom;
                """,
                (champ_no,),
            )
            return [dict(e) for e in cur.fetchall()]
    except psycopg2.Error as e:
        app.logger.error(f"Erreur DAO get_enseignants_par_champ pour {champ_no}: {e}")
        if db and not db.closed:
            db.rollback()
        return []


def get_enseignant_champ_no(enseignant_id: int) -> str | None:
    """
    Récupère le numéro de champ auquel un enseignant est associé.
    """
    db = get_db()
    if not db:
        return None
    try:
        with db.cursor() as cur:
            cur.execute("SELECT ChampNo FROM Enseignants WHERE EnseignantID = %s;", (enseignant_id,))
            result = cur.fetchone()
            return result[0] if result else None
    except psycopg2.Error as e:
        app.logger.error(f"Erreur DAO get_enseignant_champ_no pour {enseignant_id}: {e}")
        return None


def get_all_enseignants_avec_details() -> list[dict]:
    """
    Récupère tous les enseignants de la base de données avec des détails enrichis :
    informations sur leur champ (y compris le verrouillage), leurs périodes de cours et autres.
    Les enseignants sont triés par numéro de champ, puis par statut fictif, puis par nom et prénom.
    Cette fonction est optimisée pour récupérer toutes les attributions en une seule requête.
    """
    db = get_db()
    if not db:
        return []
    enseignants_complets = []
    try:
        with db.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            cur.execute(
                """
                SELECT e.EnseignantID, e.NomComplet, e.Nom, e.Prenom, e.EstTempsPlein, e.EstFictif,
                       e.ChampNo, ch.ChampNom, e.PeutChoisirHorsChampPrincipal, ch.EstVerrouille
                FROM Enseignants e JOIN Champs ch ON e.ChampNo = ch.ChampNo
                ORDER BY e.ChampNo, e.EstFictif, e.Nom, e.Prenom;
                """
            )
            enseignants_bruts = [dict(row) for row in cur.fetchall()]

        toutes_les_attributions = get_toutes_les_attributions()
        attributions_par_enseignant = {}
        for attr in toutes_les_attributions:
            attributions_par_enseignant.setdefault(attr["enseignantid"], []).append(attr)

        for ens_brut in enseignants_bruts:
            attributions_de_l_enseignant = attributions_par_enseignant.get(ens_brut["enseignantid"], [])
            periodes = calculer_periodes_pour_attributions(attributions_de_l_enseignant)
            compte_pour_moyenne_champ = ens_brut["esttempsplein"] and not ens_brut["estfictif"]
            enseignants_complets.append(
                {
                    **ens_brut,
                    "periodes_cours": periodes["periodes_cours"],
                    "periodes_autres": periodes["periodes_autres"],
                    "total_periodes": periodes["total_periodes"],
                    "compte_pour_moyenne_champ": compte_pour_moyenne_champ,
                }
            )
        return enseignants_complets
    except psycopg2.Error as e:
        app.logger.error(f"Erreur DAO get_all_enseignants_avec_details: {e}")
        if db and not db.closed:
            db.rollback()
        return []


def get_cours_disponibles_par_champ(champ_no: str) -> list[dict]:
    """
    Récupère les cours disponibles pour un champ donné, en calculant les groupes restants.
    Retourne une liste de dictionnaires, triés par type de cours (autre ou non) puis par code de cours.
    """
    db = get_db()
    if not db:
        return []
    try:
        with db.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            cur.execute(
                """
                SELECT c.CodeCours, c.CoursDescriptif, c.NbPeriodes, c.EstCoursAutre, c.NbGroupeInitial,
                       (c.NbGroupeInitial - COALESCE(SUM(ac.NbGroupesPris), 0)) AS grprestant
                FROM Cours c LEFT JOIN AttributionsCours ac ON c.CodeCours = ac.CodeCours
                WHERE c.ChampNo = %s
                GROUP BY c.CodeCours, c.CoursDescriptif, c.NbPeriodes, c.EstCoursAutre, c.NbGroupeInitial
                ORDER BY c.EstCoursAutre, c.CodeCours;
                """,
                (champ_no,),
            )
            return [dict(cr) for cr in cur.fetchall()]
    except psycopg2.Error as e:
        app.logger.error(f"Erreur DAO get_cours_disponibles_par_champ pour {champ_no}: {e}")
        if db and not db.closed:
            db.rollback()
        return []


def get_attributions_enseignant(enseignant_id: int) -> list[dict]:
    """
    Récupère toutes les attributions de cours pour un enseignant spécifique.
    Inclut les détails du cours (descriptif, périodes, type, champ d'origine).
    Retourne une liste de dictionnaires.
    """
    db = get_db()
    if not db:
        return []
    try:
        with db.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            cur.execute(
                """
                SELECT ac.AttributionID, ac.CodeCours, ac.NbGroupesPris, c.CoursDescriptif,
                       c.NbPeriodes, c.EstCoursAutre, c.ChampNo AS ChampOrigineCours
                FROM AttributionsCours ac JOIN Cours c ON ac.CodeCours = c.CodeCours
                WHERE ac.EnseignantID = %s
                ORDER BY c.EstCoursAutre, c.CoursDescriptif;
                """,
                (enseignant_id,),
            )
            return [dict(a) for a in cur.fetchall()]
    except psycopg2.Error as e:
        app.logger.error(f"Erreur DAO get_attributions_enseignant pour {enseignant_id}: {e}")
        if db and not db.closed:
            db.rollback()
        return []


def get_toutes_les_attributions() -> list[dict]:
    """
    Récupère toutes les attributions de cours pour tous les enseignants en une seule requête.
    Cette fonction est utilisée pour optimiser les performances en évitant de multiples requêtes
    lors du calcul des totaux pour tous les enseignants.
    """
    db = get_db()
    if not db:
        return []
    try:
        with db.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            cur.execute(
                """
                SELECT ac.AttributionID, ac.EnseignantID, ac.CodeCours, ac.NbGroupesPris,
                       c.CoursDescriptif, c.NbPeriodes, c.EstCoursAutre
                FROM AttributionsCours ac JOIN Cours c ON ac.CodeCours = c.CodeCours
                ORDER BY ac.EnseignantID, c.EstCoursAutre, c.CoursDescriptif;
                """
            )
            return [dict(a) for a in cur.fetchall()]
    except psycopg2.Error as e:
        app.logger.error(f"Erreur DAO get_toutes_les_attributions: {e}")
        if db and not db.closed:
            db.rollback()
        return []


def calculer_periodes_pour_attributions(attributions: list[dict]) -> dict[str, float]:
    """
    Calcule le total des périodes de cours (enseignement) et des périodes d'autres tâches
    à partir d'une liste d'attributions fournie.
    Cette fonction est générique et peut être utilisée avec des listes d'attributions déjà chargées en mémoire.
    Utilise des floats pour les calculs de périodes.
    """
    periodes_enseignement = sum(float(a["nbperiodes"]) * a["nbgroupespris"] for a in attributions if not a["estcoursautre"])
    periodes_autres = sum(float(a["nbperiodes"]) * a["nbgroupespris"] for a in attributions if a["estcoursautre"])
    return {
        "periodes_cours": periodes_enseignement,
        "periodes_autres": periodes_autres,
        "total_periodes": periodes_enseignement + periodes_autres,
    }


def calculer_periodes_enseignant(enseignant_id: int) -> dict[str, float]:
    """
    Calcule le total des périodes de cours et d'autres tâches pour un enseignant spécifique
    en interrogeant la base de données pour ses attributions.
    """
    attributions = get_attributions_enseignant(enseignant_id)
    return calculer_periodes_pour_attributions(attributions)


def get_groupes_restants_pour_cours(code_cours: str) -> int:
    """
    Calcule le nombre de groupes restants pour un cours spécifique.
    Prend en compte le nombre initial de groupes et le nombre de groupes déjà attribués.
    """
    db = get_db()
    if not db:
        app.logger.warning(f"get_groupes_restants_pour_cours: Connexion DB échouée pour {code_cours}.")
        return 0
    try:
        with db.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            cur.execute(
                """
                SELECT (c.NbGroupeInitial - COALESCE(SUM(ac.NbGroupesPris), 0)) AS grp_dispo
                FROM Cours c
                LEFT JOIN AttributionsCours ac ON c.CodeCours = ac.CodeCours AND c.CodeCours = %s
                WHERE c.CodeCours = %s
                GROUP BY c.NbGroupeInitial, c.CodeCours;
                """,
                (code_cours, code_cours),
            )
            result = cur.fetchone()

            if result is None or result["grp_dispo"] is None:
                # Si le cours n'a aucune attribution ou la requête ne renvoie rien,
                # on récupère directement le nombre initial de groupes du cours.
                cur.execute("SELECT NbGroupeInitial FROM Cours WHERE CodeCours = %s", (code_cours,))
                initial_groups_info = cur.fetchone()
                return initial_groups_info["nbgroupeinitial"] if initial_groups_info else 0
            return result["grp_dispo"]
    except psycopg2.Error as e:
        app.logger.error(f"Erreur DAO get_groupes_restants_pour_cours pour {code_cours}: {e}")
        if db and not db.closed:
            db.rollback()
        return 0


def get_all_cours_avec_details_champ() -> list[dict]:
    """
    Récupère tous les cours disponibles dans la base de données avec les détails de leur champ d'origine.
    Trié par code de cours.
    """
    db = get_db()
    if not db:
        return []
    try:
        with db.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            cur.execute(
                """
                SELECT c.CodeCours, c.CoursDescriptif, c.ChampNo, ch.ChampNom
                FROM Cours c JOIN Champs ch ON c.ChampNo = ch.ChampNo ORDER BY c.CodeCours;
                """
            )
            return [dict(row) for row in cur.fetchall()]
    except psycopg2.Error as e:
        app.logger.error(f"Erreur DAO get_all_cours_avec_details_champ: {e}")
        if db and not db.closed:
            db.rollback()
        return []


def toggle_champ_lock_status(champ_no: str) -> bool | None:
    """
    Bascule le statut de verrouillage (EstVerrouille) d'un champ spécifique.
    Retourne le nouveau statut de verrouillage (True si verrouillé, False si déverrouillé)
    ou None en cas d'erreur.
    """
    db = get_db()
    if not db:
        return None
    try:
        with db.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            cur.execute(
                "UPDATE Champs SET EstVerrouille = NOT EstVerrouille WHERE ChampNo = %s RETURNING EstVerrouille;",
                (champ_no,),
            )
            result = cur.fetchone()
            db.commit()
            return result["estverrouille"] if result else None
    except psycopg2.Error as e:
        if db and not db.closed:
            db.rollback()
        app.logger.error(f"Erreur DAO toggle_champ_lock_status pour {champ_no}: {e}")
        return None


# --- ROUTES DE L'APPLICATION (Pages HTML) ---


@app.route("/login", methods=["GET", "POST"])
def login() -> Any:
    """
    Gère la connexion des utilisateurs.
    GET: Affiche le formulaire de connexion.
    POST: Traite les données de connexion.
    """
    if current_user.is_authenticated:
        flash("Vous êtes déjà connecté(e).", "info")
        return redirect(url_for("index"))

    # Déterminer si c'est le tout premier utilisateur à créer dans la base de données.
    _first_user_check = get_users_count() == 0

    if request.method == "POST":
        username = request.form["username"].strip()
        password = request.form["password"].strip()

        user_data = get_user_by_username(username)

        if user_data and check_hashed_password(user_data["password_hash"], password):
            # Lors de la connexion, recharger l'utilisateur via get_user_by_id pour obtenir ses allowed_champs
            user = get_user_by_id(user_data["id"])
            if user:
                login_user(user)
                flash(f"Connexion réussie! Bienvenue, {user.username}.", "success")
                # Redirige vers la page demandée avant la connexion, ou vers l'accueil
                next_page = request.args.get("next")
                return redirect(next_page or url_for("index"))
            flash("Erreur lors du chargement des informations utilisateur.", "error")
        else:
            flash("Nom d'utilisateur ou mot de passe invalide.", "error")
    return render_template("login.html", first_user=_first_user_check)


@app.route("/logout")
@login_required
def logout() -> Any:
    """
    Déconnecte l'utilisateur actuel.
    """
    logout_user()
    flash("Vous avez été déconnecté(e).", "info")
    return redirect(url_for("index"))


@app.route("/register", methods=["GET", "POST"])
def register() -> Any:
    """
    Gère l'inscription du tout premier utilisateur seulement.
    Après le premier utilisateur, cette route est désactivée.
    Les nouveaux utilisateurs devront être créés via le panneau d'administration.
    """
    # Déterminer si c'est le tout premier utilisateur à être créé dans la base de données.
    user_count = get_users_count()

    # Si des utilisateurs existent déjà, rediriger l'utilisateur.
    # Cette route est réservée à la création du tout premier compte ADMIN.
    if user_count > 0:
        if current_user.is_authenticated and current_user.is_admin:
            flash("Les nouveaux comptes sont créés via la page d'administration des utilisateurs.", "info")
            return redirect(url_for("page_administration_utilisateurs"))
        flash("L'inscription directe est désactivée. Veuillez contacter un administrateur.", "error")
        return redirect(url_for("login"))

    # Logique pour le tout premier utilisateur (sera ADMIN d'office)
    _username = request.form.get("username", "").strip()  # pour réafficher le nom d'utilisateur en cas d'erreur
    # _is_admin_checked = True # Le premier utilisateur est toujours admin, cette variable n'est pas vraiment utilisée.

    if request.method == "POST":
        username = request.form["username"].strip()
        password = request.form["password"].strip()
        confirm_password = request.form["confirm_password"].strip()

        if not username or not password or not confirm_password:
            flash("Tous les champs sont requis.", "error")
            return render_template("register.html", username=_username, first_user=True)

        if password != confirm_password:
            flash("Les mots de passe ne correspondent pas.", "error")
            return render_template("register.html", username=_username, first_user=True)

        if len(password) < 6:
            flash("Le mot de passe doit contenir au moins 6 caractères.", "error")
            return render_template("register.html", username=_username, first_user=True)

        user = create_user(username, password, is_admin=True)  # Le premier utilisateur est toujours admin
        if user:
            # Accordez l'accès à tous les champs au premier utilisateur admin
            grant_access_to_all_champs(user.id)
            app.logger.info(f"Premier utilisateur {user.username} créé et accès à tous les champs accordé.")

            flash(f"Compte '{username}' créé avec succès! Vous pouvez maintenant vous connecter.", "success")
            return redirect(url_for("login"))
        flash("Ce nom d'utilisateur est déjà pris.", "error")  # Erreur si create_user retourne None (nom d'utilisateur existant)

    # Pour la requête GET, passer la variable first_user au template
    return render_template("register.html", first_user=True)


@app.route("/")
@login_required  # Protection de la route
def index() -> Any:
    """
    Affiche la page d'accueil de l'application.
    Les champs affichés dépendent des permissions de l'utilisateur.
    """
    all_champs = get_all_champs()
    if current_user.is_admin:
        champs_accessible = all_champs
    else:
        # Filtrer les champs pour les utilisateurs non-admin
        champs_accessible = [champ for champ in all_champs if current_user.can_access_champ(champ["champno"])]
    return render_template("index.html", champs=champs_accessible)


@app.route("/champ/<string:champ_no>")
@login_required  # Protection de la route
def page_champ(champ_no: str) -> Any:
    """
    Affiche la page détaillée d'un champ spécifique.
    Vérifie les permissions de l'utilisateur pour le champ.
    """
    # Première vérification de permission
    if not current_user.can_access_champ(champ_no):
        flash("Vous n'avez pas les permissions pour accéder à ce champ.", "error")
        return redirect(url_for("index"))

    champ_details = get_champ_details(champ_no)
    if not champ_details:
        flash(f"Le champ {champ_no} n'a pas été trouvé.", "error")
        return redirect(url_for("index"))

    enseignants_du_champ = get_enseignants_par_champ(champ_no)

    ids_enseignants = [e["enseignantid"] for e in enseignants_du_champ]
    toutes_les_attributions = get_toutes_les_attributions()
    attributions_par_enseignant = {}
    for attr in toutes_les_attributions:
        if attr["enseignantid"] in ids_enseignants:
            attributions_par_enseignant.setdefault(attr["enseignantid"], []).append(attr)

    enseignants_complets = []
    total_periodes_tp_pour_moyenne = 0.0
    nb_enseignants_tp_pour_moyenne = 0

    for ens in enseignants_du_champ:
        enseignant_id = ens["enseignantid"]
        attributions_de_l_enseignant = attributions_par_enseignant.get(enseignant_id, [])
        periodes = calculer_periodes_pour_attributions(attributions_de_l_enseignant)

        enseignants_complets.append(
            {
                **ens,
                "attributions": attributions_de_l_enseignant,
                "periodes_actuelles": periodes,
            }
        )

        if ens["esttempsplein"] and not ens["estfictif"]:
            total_periodes_tp_pour_moyenne += periodes["total_periodes"]
            nb_enseignants_tp_pour_moyenne += 1

    moyenne_champ = (total_periodes_tp_pour_moyenne / nb_enseignants_tp_pour_moyenne) if nb_enseignants_tp_pour_moyenne > 0 else 0.0

    cours_disponibles_bruts = get_cours_disponibles_par_champ(champ_no)
    cours_enseignement_champ = [c for c in cours_disponibles_bruts if not c["estcoursautre"]]
    cours_autres_taches_champ = [c for c in cours_disponibles_bruts if c["estcoursautre"]]

    return render_template(
        "page_champ.html",
        champ=champ_details,
        enseignants_data=enseignants_complets,
        cours_enseignement_champ=cours_enseignement_champ,
        cours_autres_taches_champ=cours_autres_taches_champ,
        moyenne_champ_initiale=moyenne_champ,
    )


@app.route("/sommaire")
@login_required  # Protection de la route
def page_sommaire() -> Any:
    """
    Affiche la page du sommaire global des tâches des enseignants.
    """
    enseignants_par_champ_data, moyennes_champs, moyenne_gen = calculer_donnees_sommaire()
    return render_template(
        "page_sommaire.html",
        enseignants_par_champ=enseignants_par_champ_data,
        moyennes_par_champ=moyennes_champs,
        moyenne_generale=moyenne_gen,
    )


@app.route("/administration")
@login_required  # Requiert une connexion
@admin_required  # Requiert des droits d'administrateur
def page_administration_donnees() -> Any:
    """
    Affiche la page d'administration générale (import/export).
    """
    cours_pour_reassignation = get_all_cours_avec_details_champ()
    champs_pour_destination = get_all_champs()
    return render_template(
        "administration_donnees.html",
        cours_a_reassigner=cours_pour_reassignation or [],
        champs_destination=champs_pour_destination or [],
    )


@app.route("/administration/utilisateurs")
@login_required
@admin_required
def page_administration_utilisateurs() -> Any:
    """
    Affiche la page d'administration des utilisateurs et de leurs permissions d'accès aux champs.
    """
    users_with_access = get_all_users_with_access_info()
    all_champs = get_all_champs()
    return render_template("administration_utilisateurs.html", users=users_with_access, all_champs=all_champs)


# --- Fonctions utilitaires pour le sommaire ---
def calculer_donnees_sommaire() -> tuple[list[dict], dict[str, dict], float]:
    """
    Calcule les données agrégées nécessaires pour la page sommaire globale.
    """
    tous_enseignants_details = get_all_enseignants_avec_details()
    enseignants_par_champ_temp: dict[str, Any] = {}
    moyennes_par_champ_calculees: dict[str, dict] = {}
    total_periodes_global_tp = 0.0
    nb_enseignants_tp_global = 0

    for ens in tous_enseignants_details:
        champ_no = ens["champno"]
        champ_nom = ens["champnom"]
        est_verrouille = ens["estverrouille"]

        if champ_no not in enseignants_par_champ_temp:
            enseignants_par_champ_temp[champ_no] = {
                "champno": champ_no,
                "champnom": champ_nom,
                "enseignants": [],
                "total_periodes_cours_champ": 0.0,
                "total_periodes_autres_champ": 0.0,
                "total_periodes_champ": 0.0,
            }

        enseignants_par_champ_temp[champ_no]["enseignants"].append(ens)
        enseignants_par_champ_temp[champ_no]["total_periodes_cours_champ"] += ens["periodes_cours"]
        enseignants_par_champ_temp[champ_no]["total_periodes_autres_champ"] += ens["periodes_autres"]
        enseignants_par_champ_temp[champ_no]["total_periodes_champ"] += ens["total_periodes"]

        if ens["compte_pour_moyenne_champ"]:
            if champ_no not in moyennes_par_champ_calculees:
                moyennes_par_champ_calculees[champ_no] = {
                    "champ_nom": champ_nom,
                    "total_periodes": 0.0,
                    "nb_enseignants": 0,
                    "moyenne": 0.0,
                    "est_verrouille": est_verrouille,
                }
            moyennes_par_champ_calculees[champ_no]["total_periodes"] += ens["total_periodes"]
            moyennes_par_champ_calculees[champ_no]["nb_enseignants"] += 1

            total_periodes_global_tp += ens["total_periodes"]
            nb_enseignants_tp_global += 1

    for data_champ in moyennes_par_champ_calculees.values():
        if data_champ["nb_enseignants"] > 0:
            data_champ["moyenne"] = data_champ["total_periodes"] / data_champ["nb_enseignants"]
        del data_champ["total_periodes"]
        del data_champ["nb_enseignants"]

    moyenne_generale_calculee = (
        (total_periodes_global_tp / nb_enseignants_tp_global) if nb_enseignants_tp_global > 0 else 0.0
    )
    enseignants_par_champ_final = list(enseignants_par_champ_temp.values())

    return enseignants_par_champ_final, moyennes_par_champ_calculees, moyenne_generale_calculee


# --- API ENDPOINTS ---
@app.route("/api/sommaire/donnees", methods=["GET"])
@login_required  # Protection de la route API
def api_get_donnees_sommaire() -> Any:
    """
    API pour récupérer les données actualisées du sommaire global de l'établissement.
    """
    enseignants_groupes, moyennes_champs, moyenne_gen = calculer_donnees_sommaire()
    return jsonify(
        {
            "enseignants_par_champ": enseignants_groupes,
            "moyennes_par_champ": moyennes_champs,
            "moyenne_generale": moyenne_gen,
        }
    )


@app.route("/api/attributions/ajouter", methods=["POST"])
@login_required  # Protection de la route API
def api_ajouter_attribution() -> Any:
    """
    API pour ajouter une attribution de cours à un enseignant.
    Vérifie les permissions de l'utilisateur sur le champ et le statut de verrouillage.
    """
    db = get_db()
    if not db:
        return jsonify({"success": False, "message": "Erreur de connexion à la base de données."}), 500

    data = request.get_json()
    if not data:
        return jsonify({"success": False, "message": "Aucune donnée JSON reçue."}), 400

    enseignant_id = data.get("enseignant_id")
    code_cours = data.get("code_cours")
    if not enseignant_id or not code_cours:
        return jsonify({"success": False, "message": "Données manquantes (ID enseignant ou code cours)."}), 400

    champ_no_enseignant = get_enseignant_champ_no(enseignant_id)
    if not champ_no_enseignant:
        return jsonify({"success": False, "message": "Enseignant ou son champ non trouvé."}), 404

    # Vérification d'accès au champ par l'utilisateur connecté
    if not current_user.can_access_champ(champ_no_enseignant):
        return jsonify({"success": False, "message": "Vous n'avez pas les permissions pour attribuer des cours dans ce champ."}), 403

    nouvelle_attribution_id = None
    periodes_enseignant_maj: dict[str, float] = {}
    groupes_restants_cours_maj = 0
    attributions_enseignant_maj: list[dict] = []
    infos_enseignant_pour_reponse: dict[str, Any] = {}

    try:
        with db.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            # Vérifier le statut de verrouillage du champ de l'enseignant.
            # Les enseignants fictifs (tâches restantes) sont exemptés du verrouillage.
            query_verrou = """
                SELECT e.EstFictif, ch.EstVerrouille
                FROM Enseignants e JOIN Champs ch ON e.ChampNo = ch.ChampNo
                WHERE e.EnseignantID = %s AND ch.ChampNo = %s;
            """
            cur.execute(query_verrou, (enseignant_id, champ_no_enseignant))
            verrou_info = cur.fetchone()

            if verrou_info and verrou_info["estverrouille"] and not verrou_info["estfictif"]:
                msg = "Impossible d'attribuer un cours, le champ est verrouillé. " "Seules les tâches restantes peuvent être modifiées."
                return jsonify({"success": False, "message": msg}), 403

            # Vérifier les groupes disponibles pour le cours avant l'insertion.
            query_grp_dispo = """
                SELECT (c.NbGroupeInitial - COALESCE(SUM(ac.NbGroupesPris), 0)) AS grp_dispo
                FROM Cours c
                LEFT JOIN AttributionsCours ac ON c.CodeCours = ac.CodeCours AND c.CodeCours = %s
                WHERE c.CodeCours = %s
                GROUP BY c.NbGroupeInitial, c.CodeCours;
            """
            cur.execute(query_grp_dispo, (code_cours, code_cours))
            cours_info = cur.fetchone()

            groupes_disponibles_actuels = 0
            if cours_info and cours_info["grp_dispo"] is not None:
                groupes_disponibles_actuels = cours_info["grp_dispo"]
            else:
                cur.execute("SELECT NbGroupeInitial FROM Cours WHERE CodeCours = %s", (code_cours,))
                initial_groups_info = cur.fetchone()
                if initial_groups_info:
                    groupes_disponibles_actuels = initial_groups_info["nbgroupeinitial"]
                else:
                    return jsonify({"success": False, "message": "Cours non trouvé."}), 404

            if groupes_disponibles_actuels < 1:
                return jsonify({"success": False, "message": "Plus de groupes disponibles pour ce cours."}), 409

            # Insère la nouvelle attribution (toujours 1 groupe à la fois).
            cur.execute(
                "INSERT INTO AttributionsCours (EnseignantID, CodeCours, NbGroupesPris) VALUES (%s, %s, %s) RETURNING AttributionID;",
                (enseignant_id, code_cours, 1),
            )
            resultat_insertion = cur.fetchone()
            if resultat_insertion is None:
                db.rollback()
                msg = f"Échec de l'insertion d'attribution pour ens {enseignant_id}, cours {code_cours}."
                app.logger.error(msg)
                return jsonify({"success": False, "message": "Erreur interne lors de la création de l'attribution."}), 500
            nouvelle_attribution_id = resultat_insertion["attributionid"]

            cur.execute(
                "SELECT ChampNo, EstTempsPlein, EstFictif, Nom, Prenom FROM Enseignants WHERE EnseignantID = %s",
                (enseignant_id,),
            )
            resultat_enseignant = cur.fetchone()
            if resultat_enseignant:
                infos_enseignant_pour_reponse = dict(resultat_enseignant)
            db.commit()

        try:
            periodes_enseignant_maj = calculer_periodes_enseignant(enseignant_id)
            groupes_restants_cours_maj = get_groupes_restants_pour_cours(code_cours)
            attributions_enseignant_maj = get_attributions_enseignant(enseignant_id)
        except Exception as e_calcul:  # pylint: disable=broad-except
            msg = f"Erreur post-attribution (calculs) pour ens {enseignant_id}, cours {code_cours}: {e_calcul}"
            app.logger.error(msg)
            message_succes_partiel = "Cours attribué, mais erreur lors de la mise à jour des totaux."
            return (
                jsonify(
                    {
                        "success": True,
                        "message": message_succes_partiel,
                        "attribution_id": nouvelle_attribution_id,
                        "enseignant_id": enseignant_id,
                        "code_cours": code_cours,
                        "periodes_enseignant": {},
                        "groupes_restants_cours": -1,
                        "attributions_enseignant": [],
                        **infos_enseignant_pour_reponse,
                    }
                ),
                201,
            )

        return (
            jsonify(
                {
                    "success": True,
                    "message": "Cours attribué avec succès!",
                    "attribution_id": nouvelle_attribution_id,
                    "enseignant_id": enseignant_id,
                    "code_cours": code_cours,
                    "periodes_enseignant": periodes_enseignant_maj,
                    "groupes_restants_cours": groupes_restants_cours_maj,
                    "attributions_enseignant": attributions_enseignant_maj,
                    **infos_enseignant_pour_reponse,
                }
            ),
            201,
        )

    except psycopg2.Error as e_psy:
        if db and not db.closed:
            db.rollback()
        app.logger.error(f"Erreur psycopg2 API ajouter attribution: {e_psy}")
        return jsonify({"success": False, "message": "Erreur de base de données lors de l'ajout de l'attribution."}), 500
    except Exception as e_gen:  # pylint: disable=broad-except
        if db and not db.closed:
            db.rollback()
        app.logger.error(f"Erreur générale Exception API ajouter attribution: {e_gen}")
        return jsonify({"success": False, "message": "Erreur serveur inattendue lors de l'ajout de l'attribution."}), 500


@app.route("/api/attributions/supprimer", methods=["POST"])
@login_required  # Protection de la route API
def api_supprimer_attribution() -> Any:
    """
    API pour supprimer une attribution de cours spécifique.
    Vérifie les permissions de l'utilisateur sur le champ et le statut de verrouillage.
    """
    db = get_db()
    if not db:
        return jsonify({"success": False, "message": "Erreur de connexion à la base de données."}), 500

    data = request.get_json()
    if not data:
        return jsonify({"success": False, "message": "Aucune donnée JSON reçue."}), 400
    attribution_id_a_supprimer = data.get("attribution_id")
    if not attribution_id_a_supprimer:
        return jsonify({"success": False, "message": "ID d'attribution manquant."}), 400

    enseignant_id_concerne, code_cours_concerne = None, None
    periodes_enseignant_maj: dict[str, float] = {}
    groupes_restants_cours_maj = 0
    attributions_enseignant_maj: list[dict] = []
    periodes_liberees_par_suppression = 0.0
    infos_enseignant_pour_reponse: dict[str, Any] = {}

    try:
        with db.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            # Récupérer les informations de l'attribution avant suppression et vérifier le verrouillage.
            query_select_attribution = """
                SELECT ac.EnseignantID, ac.CodeCours, c.NbPeriodes AS PeriodesDuCours, ac.NbGroupesPris,
                       e.EstFictif, ch.EstVerrouille, ch.ChampNo
                FROM AttributionsCours ac
                JOIN Cours c ON ac.CodeCours = c.CodeCours
                JOIN Enseignants e ON ac.EnseignantID = e.EnseignantID
                JOIN Champs ch ON e.ChampNo = ch.ChampNo
                WHERE ac.AttributionID = %s;
            """
            cur.execute(query_select_attribution, (attribution_id_a_supprimer,))
            attribution_info = cur.fetchone()
            if not attribution_info:
                return jsonify({"success": False, "message": "Attribution non trouvée."}), 404

            champ_no_concerne = attribution_info["champno"]
            # Vérification d'accès au champ par l'utilisateur connecté
            if not current_user.can_access_champ(champ_no_concerne):
                return jsonify({"success": False, "message": "Vous n'avez pas les permissions pour retirer des cours dans ce champ."}), 403

            if attribution_info["estverrouille"] and not attribution_info["estfictif"]:
                msg = "Impossible de retirer ce cours, le champ est verrouillé. " "Seules les tâches restantes peuvent être modifiées."
                return jsonify({"success": False, "message": msg}), 403

            enseignant_id_concerne = attribution_info["enseignantid"]
            code_cours_concerne = attribution_info["codecours"]
            periodes_liberees_par_suppression = float(attribution_info.get("periodesducours", 0.0)) * attribution_info.get(
                "nbgroupespris", 0
            )

            # Supprime l'attribution.
            cur.execute("DELETE FROM AttributionsCours WHERE AttributionID = %s;", (attribution_id_a_supprimer,))
            if cur.rowcount == 0:
                db.rollback()
                return jsonify({"success": False, "message": "Attribution non trouvée ou déjà supprimée (concurrence?)."}), 404

            cur.execute(
                "SELECT ChampNo, EstTempsPlein, EstFictif, Nom, Prenom FROM Enseignants WHERE EnseignantID = %s",
                (enseignant_id_concerne,),
            )
            resultat_enseignant = cur.fetchone()
            if resultat_enseignant:
                infos_enseignant_pour_reponse = dict(resultat_enseignant)
            db.commit()

        try:
            periodes_enseignant_maj = calculer_periodes_enseignant(enseignant_id_concerne)
            groupes_restants_cours_maj = get_groupes_restants_pour_cours(code_cours_concerne)
            attributions_enseignant_maj = get_attributions_enseignant(enseignant_id_concerne)
        except Exception as e_calcul:  # pylint: disable=broad-except
            msg = f"Erreur post-suppression (calculs) pour ens {enseignant_id_concerne}, cours {code_cours_concerne}: {e_calcul}"
            app.logger.error(msg)
            message_succes_partiel = "Attribution supprimée, mais erreur lors de la mise à jour des totaux."
            return (
                jsonify(
                    {
                        "success": True,
                        "message": message_succes_partiel,
                        "enseignant_id": enseignant_id_concerne,
                        "code_cours": code_cours_concerne,
                        "nb_periodes_cours_libere": periodes_liberees_par_suppression,
                        "periodes_enseignant": {},
                        "groupes_restants_cours": -1,
                        "attributions_enseignant": [],
                        **infos_enseignant_pour_reponse,
                    }
                ),
                200,
            )

        return (
            jsonify(
                {
                    "success": True,
                    "message": "Attribution supprimée avec succès!",
                    "enseignant_id": enseignant_id_concerne,
                    "code_cours": code_cours_concerne,
                    "periodes_enseignant": periodes_enseignant_maj,
                    "groupes_restants_cours": groupes_restants_cours_maj,
                    "attributions_enseignant": attributions_enseignant_maj,
                    "nb_periodes_cours_libere": periodes_liberees_par_suppression,
                    **infos_enseignant_pour_reponse,
                }
            ),
            200,
        )

    except psycopg2.Error as e_psy:
        if db and not db.closed:
            db.rollback()
        app.logger.error(f"Erreur psycopg2 API supprimer attribution: {e_psy}")
        if hasattr(e_psy, "pgcode") and e_psy.pgcode == "23503":
            return jsonify({"success": False, "message": "Suppression impossible car l'attribution est référencée ailleurs."}), 409
        return jsonify({"success": False, "message": "Erreur de base de données lors de la suppression de l'attribution."}), 500
    except Exception as e_gen:  # pylint: disable=broad-except
        if db and not db.closed:
            db.rollback()
        app.logger.error(f"Erreur générale Exception API supprimer enseignant: {e_gen}")
        msg = f"Erreur serveur inattendue lors de la suppression: {str(e_gen)}"
        return jsonify({"success": False, "message": msg}), 500


@app.route("/api/champs/<string:champ_no>/taches_restantes/creer", methods=["POST"])
@login_required  # Protection de la route API
def api_creer_tache_restante(champ_no: str) -> Any:
    """
    API pour créer une nouvelle tâche restante (représentée par un enseignant fictif) dans un champ donné.
    Vérifie les permissions de l'utilisateur sur le champ.
    """
    # Vérification d'accès au champ par l'utilisateur connecté
    if not current_user.can_access_champ(champ_no):
        return jsonify({"success": False, "message": "Vous n'avez pas les permissions pour créer une tâche restante dans ce champ."}), 403

    db = get_db()
    if not db:
        return jsonify({"success": False, "message": "Erreur de connexion à la base de données."}), 500

    nouvel_enseignant_fictif_cree: dict[str, Any] = {}
    periodes_initiales_tache: dict[str, float] = {"periodes_cours": 0.0, "periodes_autres": 0.0, "total_periodes": 0.0}

    try:
        with db.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            query_select_taches_existantes = """
                SELECT NomComplet FROM Enseignants
                WHERE ChampNo = %s AND EstFictif = TRUE AND NomComplet LIKE %s;
            """
            pattern_nom_tache = f"{champ_no}-Tâche restante-%"
            cur.execute(query_select_taches_existantes, (champ_no, pattern_nom_tache))
            taches_existantes = cur.fetchall()

            max_numero_tache = 0
            for tache in taches_existantes:
                nom_tache_existante = tache["nomcomplet"]
                parts = nom_tache_existante.split("-")
                if len(parts) > 1:
                    num_part = parts[-1].strip()
                    if num_part.isdigit():
                        numero = int(num_part)
                        max_numero_tache = max(max_numero_tache, numero)

            nom_nouvelle_tache = f"{champ_no}-Tâche restante-{max_numero_tache + 1}"

            query_insert_tache = """
                INSERT INTO Enseignants (NomComplet, Nom, Prenom, ChampNo, EstTempsPlein, EstFictif, PeutChoisirHorsChampPrincipal)
                VALUES (%s, NULL, NULL, %s, TRUE, TRUE, FALSE)
                RETURNING EnseignantID, NomComplet, Nom, Prenom, EstTempsPlein, EstFictif, PeutChoisirHorsChampPrincipal, ChampNo;
            """
            cur.execute(query_insert_tache, (nom_nouvelle_tache, champ_no))
            resultat_insertion = cur.fetchone()
            if resultat_insertion is None:
                db.rollback()
                msg = f"Échec de l'insertion de la tâche restante '{nom_nouvelle_tache}'."
                app.logger.error(msg)
                return jsonify({"success": False, "message": "Erreur interne lors de la création de la tâche."}), 500
            nouvel_enseignant_fictif_cree = dict(resultat_insertion)
            db.commit()

        return (
            jsonify(
                {
                    "success": True,
                    "message": "Tâche restante créée avec succès!",
                    "enseignant": {
                        **nouvel_enseignant_fictif_cree,
                        "attributions": [],
                    },
                    "periodes_actuelles": periodes_initiales_tache,
                }
            ),
            201,
        )

    except psycopg2.Error as e_psy:
        if db and not db.closed:
            db.rollback()
        app.logger.error(f"Erreur psycopg2 API créer tâche restante: {e_psy}")
        return jsonify({"success": False, "message": "Erreur de base de données lors de la création de la tâche."}), 500
    except Exception as e_gen:  # pylint: disable=broad-except
        if db and not db.closed:
            db.rollback()
        app.logger.error(f"Erreur générale Exception API créer tâche restante: {e_gen}")
        return jsonify({"success": False, "message": "Erreur serveur inattendue lors de la création."}), 500


@app.route("/api/enseignants/<int:enseignant_id>/supprimer", methods=["POST"])
@login_required  # Protection de la route API
def api_supprimer_enseignant(enseignant_id: int) -> Any:
    """
    API pour supprimer un enseignant, principalement utilisée pour les tâches fictives.
    Vérifie les permissions de l'utilisateur sur le champ et le statut de verrouillage.
    """
    db = get_db()
    if not db:
        return jsonify({"success": False, "message": "Erreur de connexion à la base de données."}), 500

    # Récupérer le champ_no de l'enseignant avant toute suppression
    champ_no_enseignant = get_enseignant_champ_no(enseignant_id)
    if not champ_no_enseignant:
        return jsonify({"success": False, "message": "Enseignant ou son champ non trouvé."}), 404

    # Vérification d'accès au champ par l'utilisateur connecté
    if not current_user.can_access_champ(champ_no_enseignant):
        return jsonify({"success": False, "message": "Vous n'avez pas les permissions pour supprimer un enseignant de ce champ."}), 403

    cours_liberes_apres_suppression: list[dict] = []
    try:
        with db.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            query_info_enseignant = """
                SELECT e.EstFictif, ch.EstVerrouille
                FROM Enseignants e JOIN Champs ch ON e.ChampNo = ch.ChampNo
                WHERE e.EnseignantID = %s AND ch.ChampNo = %s;
            """
            cur.execute(query_info_enseignant, (enseignant_id, champ_no_enseignant))
            enseignant_info = cur.fetchone()

            if not enseignant_info:
                return jsonify({"success": False, "message": "Enseignant non trouvé."}), 404

            if enseignant_info["estverrouille"] and not enseignant_info["estfictif"]:
                msg = "Impossible de supprimer un enseignant d'un champ verrouillé."
                return jsonify({"success": False, "message": msg}), 403

            query_cours_affectes_enseignant = """
                SELECT DISTINCT ac.CodeCours, c.NbPeriodes
                FROM AttributionsCours ac JOIN Cours c ON ac.CodeCours = c.CodeCours
                WHERE ac.EnseignantID = %s;
            """
            cur.execute(query_cours_affectes_enseignant, (enseignant_id,))
            cours_affectes_avant_suppression = cur.fetchall()

            cur.execute("DELETE FROM Enseignants WHERE EnseignantID = %s;", (enseignant_id,))
            if cur.rowcount == 0:
                db.rollback()
                return jsonify({"success": False, "message": "L'enseignant n'a pas pu être supprimé (déjà supprimé?)."}), 404
            db.commit()

            if cours_affectes_avant_suppression:
                codes_cours_uniques = list(set(c["codecours"] for c in cours_affectes_avant_suppression))
                for code_cours_unique in codes_cours_uniques:
                    groupes_restants_maj = get_groupes_restants_pour_cours(code_cours_unique)
                    nb_periodes = float(
                        next((c["nbperiodes"] for c in cours_affectes_avant_suppression if c["codecours"] == code_cours_unique), 0.0)
                    )
                    cours_liberes_apres_suppression.append(
                        {
                            "code_cours": code_cours_unique,
                            "nouveaux_groupes_restants": groupes_restants_maj,
                            "nb_periodes": nb_periodes,
                        }
                    )
        return (
            jsonify(
                {
                    "success": True,
                    "message": "Enseignant et ses attributions supprimés avec succès. Groupes de cours mis à jour.",
                    "enseignant_id": enseignant_id,
                    "cours_liberes_details": cours_liberes_apres_suppression,
                }
            ),
            200,
        )

    except psycopg2.Error as e_psy:
        if db and not db.closed:
            db.rollback()
        if hasattr(e_psy, "pgcode") and e_psy.pgcode == "23503":
            msg_err_fk = f"Suppression de l'enseignant {enseignant_id} impossible car il est référencé dans d'autres tables."
            app.logger.error(f"{msg_err_fk} Détail: {e_psy}")
            return jsonify({"success": False, "message": msg_err_fk}), 409
        app.logger.error(f"Erreur psycopg2 API supprimer enseignant: {e_psy}")
        return jsonify({"success": False, "message": "Erreur de base de données lors de la suppression de l'enseignant."}), 500
    except Exception as e_gen:  # pylint: disable=broad-except
        if db and not db.closed:
            db.rollback()
        app.logger.error(f"Erreur générale Exception API supprimer enseignant: {e_gen}")
        msg = f"Erreur serveur inattendue lors de la suppression: {str(e_gen)}"
        return jsonify({"success": False, "message": msg}), 500


@app.route("/api/champs/<string:champ_no>/basculer_verrou", methods=["POST"])
@login_required  # Protection de la route API
@admin_required  # Requiert des droits d'administrateur
def api_basculer_verrou_champ(champ_no: str) -> Any:
    """
    API pour basculer le statut de verrouillage d'un champ.
    """
    db = get_db()
    if not db:
        return jsonify({"success": False, "message": "Erreur de connexion à la base de données."}), 500

    nouveau_statut = toggle_champ_lock_status(champ_no)
    if nouveau_statut is None:
        return jsonify({"success": False, "message": f"Impossible de modifier le verrou du champ {champ_no}."}), 500

    message = f"Le champ {champ_no} a été {'verrouillé' if nouveau_statut else 'déverrouillé'}."
    return jsonify({"success": True, "message": message, "est_verrouille": nouveau_statut}), 200


@app.route("/api/utilisateurs", methods=["GET"])
@login_required
@admin_required
def api_get_all_users() -> Any:
    """
    Endpoint API pour récupérer la liste de tous les utilisateurs et le nombre d'administrateurs.
    Utilisé par la page d'administration des utilisateurs pour le rafraîchissement dynamique.
    """
    users_info = get_all_users_with_access_info()
    db = get_db()
    admin_count = 0
    if db:
        with db.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM Users WHERE is_admin = TRUE;")
            admin_count = cur.fetchone()[0]
    return jsonify({"users": users_info, "admin_count": admin_count})


@app.route("/api/utilisateurs/creer", methods=["POST"])
@login_required
@admin_required
def api_create_user() -> Any:
    """
    Endpoint API pour créer un nouvel utilisateur depuis le panneau d'administration.
    """
    data = request.get_json()
    if not data:
        return jsonify({"success": False, "message": "Aucune donnée JSON reçue."}), 400

    username = data.get("username", "").strip()
    password = data.get("password", "").strip()
    is_admin = data.get("is_admin", False)
    # Les champs autorisés sont une liste de chaînes (ChampNo)
    allowed_champs = data.get("allowed_champs", [])

    if not username or not password:
        return jsonify({"success": False, "message": "Le nom d'utilisateur et le mot de passe sont requis."}), 400
    if len(password) < 6:
        return jsonify({"success": False, "message": "Le mot de passe doit contenir au moins 6 caractères."}), 400
    if not isinstance(allowed_champs, list):
        return jsonify({"success": False, "message": "Le format des champs autorisés est invalide."}), 400

    user = create_user(username, password, is_admin)
    if user:
        if not is_admin and allowed_champs:
            # Assigner les permissions de champ si ce n'est pas un admin
            if not update_user_champ_access(user.id, allowed_champs):
                # En cas d'échec de l'attribution des permissions, supprimer l'utilisateur créé
                delete_user_data(user.id)
                app.logger.error(f"Erreur d'attribution des champs pour le nouvel utilisateur {username}.")
                return jsonify({"success": False, "message": "Utilisateur créé mais erreur lors de l'attribution des accès aux champs."}), 500

        return jsonify({"success": True, "message": f"Utilisateur '{username}' créé avec succès!", "user_id": user.id}), 201
    return jsonify({"success": False, "message": "Ce nom d'utilisateur est déjà pris."}), 409


@app.route("/api/utilisateurs/<int:user_id>/update_access", methods=["POST"])
@login_required
@admin_required
def api_update_user_access(user_id: int) -> Any:
    """
    Endpoint API pour mettre à jour les accès aux champs d'un utilisateur.
    Requiert des droits d'administrateur.
    """
    data = request.get_json()
    if not data or "champ_nos" not in data:
        return jsonify({"success": False, "message": "Données manquantes (champ_nos)."}), 400

    champ_nos_to_grant = data["champ_nos"]
    if not isinstance(champ_nos_to_grant, list):
        return jsonify({"success": False, "message": "Le format de 'champ_nos' est invalide (doit être une liste)."}), 400

    # Empêcher un administrateur de modifier ses propres droits d'accès aux champs
    # car les administrateurs ont accès à tous les champs par nature.
    if user_id == current_user.id and current_user.is_admin:
        return jsonify(
            {
                "success": False,
                "message": "Vous ne pouvez pas modifier vos propres droits d'accès aux champs si vous êtes administrateur. Les administrateurs ont accès à tous les champs par défaut.",
            }
        ), 403

    # Empêcher un administrateur de modifier les droits d'accès aux champs d'un autre administrateur.
    target_user_info = get_user_by_id(user_id)
    if not target_user_info:
        return jsonify({"success": False, "message": "Utilisateur cible non trouvé."}), 404
    if target_user_info.is_admin:
        return jsonify(
            {
                "success": False,
                "message": "Vous ne pouvez pas modifier les droits d'accès aux champs d'un autre administrateur. Les administrateurs ont accès à tous les champs par défaut.",
            }
        ), 403

    if update_user_champ_access(user_id, champ_nos_to_grant):
        return jsonify({"success": True, "message": "Accès aux champs mis à jour avec succès."}), 200
    return jsonify({"success": False, "message": "Erreur lors de la mise à jour des accès aux champs."}), 500


@app.route("/api/utilisateurs/<int:user_id>/delete", methods=["POST"])
@login_required
@admin_required
def api_delete_user(user_id: int) -> Any:
    """
    Endpoint API pour supprimer un utilisateur.
    Requiert des droits d'administrateur.
    Empêche la suppression du compte de l'utilisateur actuellement connecté
    et la suppression du dernier compte administrateur.
    """
    if user_id == current_user.id:
        return jsonify({"success": False, "message": "Vous ne pouvez pas supprimer votre propre compte."}), 403

    target_user_info = get_user_by_id(user_id)
    if not target_user_info:
        return jsonify({"success": False, "message": "Utilisateur cible non trouvé."}), 404

    # Si l'utilisateur cible est un administrateur, vérifier qu'il ne s'agit pas du dernier admin.
    if target_user_info.is_admin:
        db = get_db()
        if not db:
            return jsonify({"success": False, "message": "Erreur de connexion à la base de données."}), 500
        try:
            with db.cursor() as cur:
                cur.execute("SELECT COUNT(*) FROM Users WHERE is_admin = TRUE;")
                admin_count = cur.fetchone()[0]
                # Si il ne reste qu'un seul admin ET que l'utilisateur cible est cet admin
                if admin_count <= 1 and target_user_info.is_admin:
                    return jsonify({"success": False, "message": "Impossible de supprimer le dernier compte administrateur."}), 403
        except psycopg2.Error as e:
            app.logger.error(f"Erreur vérification admin_count: {e}")
            return jsonify({"success": False, "message": "Erreur de base de données lors de la vérification des administrateurs."}), 500

    if delete_user_data(user_id):
        return jsonify({"success": True, "message": "Utilisateur supprimé avec succès."}), 200
    return jsonify({"success": False, "message": "Erreur lors de la suppression de l'utilisateur."}), 500


# --- Fonctions utilitaires et routes pour l'importation de données Excel ---
def allowed_file(filename: str) -> bool:
    """Vérifie si l'extension du fichier est autorisée pour l'importation."""
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route("/administration/importer_cours_excel", methods=["POST"])
@login_required  # Protection de la route
@admin_required  # Requiert des droits d'administrateur
def api_importer_cours_excel() -> Any:
    """
    Importe les données des cours depuis un fichier Excel.
    """
    db = get_db()
    if not db:
        flash("Erreur de connexion à la base de données. Importation des cours annulée.", "error")
        return redirect(url_for("page_administration_donnees"))

    if "fichier_cours" not in request.files:
        flash("Aucun fichier sélectionné pour l'importation des cours.", "warning")
        return redirect(url_for("page_administration_donnees"))

    file = request.files["fichier_cours"]
    if not file or not file.filename:
        flash("Nom de fichier vide pour l'importation des cours.", "warning")
        return redirect(url_for("page_administration_donnees"))

    if not allowed_file(file.filename):
        flash("Type de fichier non autorisé pour les cours. Seuls les fichiers .xlsx sont acceptés.", "error")
        return redirect(url_for("page_administration_donnees"))

    nouveaux_cours_a_importer: list[dict] = []
    try:
        workbook = openpyxl.load_workbook(file.stream)
        sheet = workbook.active
        if not isinstance(sheet, Worksheet):
            flash("La feuille active du fichier Excel des cours est invalide.", "error")
            return redirect(url_for("page_administration_donnees"))

        iter_rows = iter(sheet.rows)
        try:
            next(iter_rows)  # Ignorer la première ligne (en-tête de colonne).
        except StopIteration:
            flash("Le fichier Excel des cours est vide ou ne contient que l'en-tête.", "warning")
            return redirect(url_for("page_administration_donnees"))

        lignes_ignorees_compt = 0
        for row_idx, row in enumerate(iter_rows, start=2):
            try:
                champ_no_raw = row[0].value
                code_cours_raw = row[1].value
                # row[2].value est la grille (ignorée)
                cours_descriptif_raw = row[3].value
                nb_groupe_initial_raw = row[4].value
                nb_periodes_raw = row[5].value
                # row[6].value est le total période (ignoré)
                est_cours_autre_raw = row[7].value

                champ_no = str(champ_no_raw).strip() if champ_no_raw is not None else None
                code_cours = str(code_cours_raw).strip() if code_cours_raw is not None else None
                cours_descriptif = str(cours_descriptif_raw).strip() if cours_descriptif_raw is not None else None

                if not all([champ_no, code_cours, cours_descriptif]):
                    flash(
                        f"Ligne {row_idx} (Cours): Données essentielles (ChampNo, CodeCours, Descriptif) manquantes. Ligne ignorée.",
                        "warning",
                    )
                    lignes_ignorees_compt += 1
                    continue

                nb_groupe_initial = 0
                if isinstance(nb_groupe_initial_raw, int | float):
                    nb_groupe_initial = int(nb_groupe_initial_raw)
                elif isinstance(nb_groupe_initial_raw, str) and nb_groupe_initial_raw.strip().isdigit():
                    nb_groupe_initial = int(nb_groupe_initial_raw.strip())
                elif nb_groupe_initial_raw is not None:
                    flash(
                        f"Ligne {row_idx} (Cours): Valeur non numérique ou incorrecte pour 'Nb Groupes Prévus' ('{nb_groupe_initial_raw}'). "
                        "Ligne ignorée.",
                        "warning",
                    )
                    lignes_ignorees_compt += 1
                    continue

                nb_periodes = 0.0
                if isinstance(nb_periodes_raw, int | float):
                    nb_periodes = float(nb_periodes_raw)
                elif isinstance(nb_periodes_raw, str):
                    try:
                        nb_periodes = float(nb_periodes_raw.replace(",", ".").strip())
                    except ValueError:
                        flash(
                            f"Ligne {row_idx} (Cours): Valeur non numérique ou incorrecte pour 'Périodes/GR' ('{nb_periodes_raw}'). Ligne ignorée.",
                            "warning",
                        )
                        lignes_ignorees_compt += 1
                        continue
                elif nb_periodes_raw is not None:
                    flash(
                        f"Ligne {row_idx} (Cours): Valeur inattendue pour 'Périodes/GR' ('{nb_periodes_raw}'). Assumée 0.0.",
                        "warning",
                    )
                    nb_periodes = 0.0

                est_cours_autre = False
                if isinstance(est_cours_autre_raw, bool):
                    est_cours_autre = est_cours_autre_raw
                elif isinstance(est_cours_autre_raw, str):
                    val_str_upper = est_cours_autre_raw.strip().upper()
                    if val_str_upper in ("VRAI", "TRUE", "1", "OUI", "YES"):
                        est_cours_autre = True
                    elif val_str_upper in ("FAUX", "FALSE", "0", "NON", "NO", ""):
                        est_cours_autre = False
                    else:
                        flash(f"Ligne {row_idx} (Cours): Valeur inattendue pour 'EstCoursAutre' ('{est_cours_autre_raw}'). Assumée FAUX.", "warning")
                elif est_cours_autre_raw is not None:
                    flash(f"Ligne {row_idx} (Cours): Valeur inattendue pour 'EstCoursAutre' ('{est_cours_autre_raw}'). Assumée FAUX.", "warning")

                nouveaux_cours_a_importer.append(
                    {
                        "codecours": code_cours,
                        "champno": champ_no,
                        "coursdescriptif": cours_descriptif,
                        "nbperiodes": nb_periodes,
                        "nbgroupeinitial": nb_groupe_initial,
                        "estcoursautre": est_cours_autre,
                    }
                )
            except IndexError:
                flash(f"Ligne {row_idx} (Cours): Nombre de colonnes insuffisant. La ligne est ignorée.", "warning")
                lignes_ignorees_compt += 1
                continue
            except (TypeError, ValueError) as te:
                flash(f"Ligne {row_idx} (Cours): Erreur de type de donnée ou de valeur. Ligne ignorée. Détails: {te}", "warning")
                lignes_ignorees_compt += 1
                continue

        if lignes_ignorees_compt > 0:
            flash(f"{lignes_ignorees_compt} ligne(s) du fichier cours ont été ignorée(s) en raison d'erreurs.", "info")

        if not nouveaux_cours_a_importer:
            flash("Aucun cours valide n'a été lu depuis le fichier. Aucune modification n'a été apportée à la base de données.", "warning")
            return redirect(url_for("page_administration_donnees"))

        with db.cursor() as cur:
            cur.execute("DELETE FROM AttributionsCours;")
            cur.execute("DELETE FROM Cours;")

            cours_importes_count = 0
            for cours_data in nouveaux_cours_a_importer:
                try:
                    cur.execute(
                        """INSERT INTO Cours (CodeCours, ChampNo, CoursDescriptif, NbPeriodes, NbGroupeInitial, EstCoursAutre)
                           VALUES (%(codecours)s, %(champno)s, %(coursdescriptif)s, %(nbperiodes)s, %(nbgroupeinitial)s, %(estcoursautre)s);""",
                        cours_data,
                    )
                    cours_importes_count += 1
                except psycopg2.Error as e_insert:
                    db.rollback()
                    err_details = e_insert.pgerror if hasattr(e_insert, "pgerror") else str(e_insert)
                    code_prob = cours_data.get("codecours", "INCONNU")
                    err_msg = (
                        f"Erreur lors de l'insertion du cours '{code_prob}' (ChampNo: {cours_data.get('champno')}): {err_details}. "
                        "L'importation des cours a été annulée. Aucune donnée n'a été modifiée."
                    )
                    flash(err_msg, "error")
                    app.logger.error(err_msg)
                    return redirect(url_for("page_administration_donnees"))
            db.commit()
            msg_succes = (
                f"{cours_importes_count} cours ont été importés avec succès. "
                "Les anciens cours et toutes les attributions existantes ont été supprimés."
            )
            flash(msg_succes, "success")

    except InvalidFileException:
        flash("Le fichier Excel des cours fourni est invalide ou corrompu.", "error")
    except AssertionError as ae:
        flash(f"Erreur de format ou de structure dans le fichier Excel des cours: {ae}", "error")
        app.logger.error(f"Erreur d'assertion lors de l'importation des cours: {ae}")
    except Exception as e:  # pylint: disable=broad-except
        if db and not db.closed and not db.autocommit:
            db.rollback()
        app.logger.error(f"Erreur inattendue lors de l'importation des cours: {type(e).__name__} - {e}")
        msg_err_inatt = f"Une erreur inattendue est survenue lors de l'importation des cours: {type(e).__name__}. L'opération a été annulée."
        flash(msg_err_inatt, "error")
    return redirect(url_for("page_administration_donnees"))


@app.route("/administration/importer_enseignants_excel", methods=["POST"])
@login_required  # Protection de la route
@admin_required  # Requiert des droits d'administrateur
def api_importer_enseignants_excel() -> Any:
    """
    Importe les données des enseignants depuis un fichier Excel.
    """
    db = get_db()
    if not db:
        flash("Erreur de connexion à la base de données. Importation des enseignants annulée.", "error")
        return redirect(url_for("page_administration_donnees"))

    if "fichier_enseignants" not in request.files:
        flash("Aucun fichier sélectionné pour l'importation des enseignants.", "warning")
        return redirect(url_for("page_administration_donnees"))

    file = request.files["fichier_enseignants"]
    if not file or not file.filename:
        flash("Nom de fichier vide pour l'importation des enseignants.", "warning")
        return redirect(url_for("page_administration_donnees"))

    if not allowed_file(file.filename):
        flash("Type de fichier non autorisé pour les enseignants. Seuls les fichiers .xlsx sont acceptés.", "error")
        return redirect(url_for("page_administration_donnees"))

    nouveaux_enseignants_a_importer: list[dict] = []
    try:
        workbook = openpyxl.load_workbook(file.stream)
        sheet = workbook.active
        if not isinstance(sheet, Worksheet):
            flash("La feuille active du fichier Excel des enseignants est invalide.", "error")
            return redirect(url_for("page_administration_donnees"))

        iter_rows = iter(sheet.rows)
        try:
            next(iter_rows)  # Ignorer la première ligne (en-tête de colonne).
        except StopIteration:
            flash("Le fichier Excel des enseignants est vide ou ne contient que l'en-tête.", "warning")
            return redirect(url_for("page_administration_donnees"))

        lignes_ignorees_compt = 0
        for row_idx, row in enumerate(iter_rows, start=2):
            try:
                champ_no_raw = row[0].value
                # row[1].value est le "Champ Nom" (ignoré)
                nom_famille_raw = row[2].value
                prenom_raw = row[3].value
                temps_plein_raw = row[4].value

                champ_no = str(champ_no_raw).strip() if champ_no_raw is not None else None
                nom_famille = str(nom_famille_raw).strip() if nom_famille_raw is not None else None
                prenom = str(prenom_raw).strip() if prenom_raw is not None else None

                if not all([champ_no, nom_famille, prenom]):
                    msg = f"Ligne {row_idx} (Enseignants): Données essentielles (ChampNo, Nom de famille, Prénom) manquantes ou vides. Ligne ignorée."
                    flash(msg, "warning")
                    lignes_ignorees_compt += 1
                    continue

                nom_complet = f"{prenom} {nom_famille}"

                est_temps_plein = True
                if isinstance(temps_plein_raw, bool):
                    est_temps_plein = temps_plein_raw
                elif isinstance(temps_plein_raw, str):
                    val_str_upper = temps_plein_raw.strip().upper()
                    if val_str_upper in ("VRAI", "TRUE", "1", "OUI", "YES"):
                        est_temps_plein = True
                    elif val_str_upper in ("FAUX", "FALSE", "0", "NON", "NO", ""):
                        est_temps_plein = False
                    else:
                        flash(f"Ligne {row_idx} (Ens): Valeur inattendue pour 'Temps Plein' ('{temps_plein_raw}'). Assumée VRAI.", "warning")
                elif temps_plein_raw is None:
                    est_temps_plein = True
                else:
                    flash(f"Ligne {row_idx} (Ens): Type de donnée inattendu pour 'Temps Plein' ('{temps_plein_raw}'). Assumée VRAI.", "warning")

                nouveaux_enseignants_a_importer.append(
                    {
                        "nomcomplet": nom_complet,
                        "nom": nom_famille,
                        "prenom": prenom,
                        "champno": champ_no,
                        "esttempsplein": est_temps_plein,
                        "estfictif": False,
                        "peutchoisirhorschampprincipal": False,
                    }
                )
            except IndexError:
                flash(f"Ligne {row_idx} (Enseignants): Nombre de colonnes insuffisant. La ligne est ignorée.", "warning")
                lignes_ignorees_compt += 1
                continue
            except (TypeError, ValueError) as te:
                flash(f"Ligne {row_idx} (Enseignants): Erreur de type de donnée ou de valeur. Ligne ignorée. Détails: {te}", "warning")
                lignes_ignorees_compt += 1
                continue

        if lignes_ignorees_compt > 0:
            flash(f"{lignes_ignorees_compt} ligne(s) du fichier enseignants ont été ignorée(s) en raison d'erreurs.", "info")

        if not nouveaux_enseignants_a_importer:
            flash("Aucun enseignant valide n'a été lu. Aucune modification apportée à la base de données.", "warning")
            return redirect(url_for("page_administration_donnees"))

        with db.cursor() as cur:
            cur.execute("DELETE FROM AttributionsCours;")
            cur.execute("DELETE FROM Enseignants;")

            ens_importes_count = 0
            for ens_data in nouveaux_enseignants_a_importer:
                try:
                    cur.execute(
                        """INSERT INTO Enseignants (NomComplet, Nom, Prenom, ChampNo, EstTempsPlein, EstFictif, PeutChoisirHorsChampPrincipal)
                           VALUES (%(nomcomplet)s, %(nom)s, %(prenom)s, %(champno)s, %(esttempsplein)s, %(estfictif)s, %(peutchoisirhorschampprincipal)s);""",
                        ens_data,
                    )
                    ens_importes_count += 1
                except psycopg2.Error as e_insert:
                    db.rollback()
                    err_details = e_insert.pgerror if hasattr(e_insert, "pgerror") else str(e_insert)
                    nom_prob = ens_data.get("nomcomplet", "INCONNU")
                    err_msg = (
                        f"Erreur lors de l'insertion de l'enseignant '{nom_prob}' (ChampNo: {ens_data.get('champno')}): {err_details}. "
                        "L'importation des enseignants a été annulée. Aucune donnée n'a été modifiée."
                    )
                    flash(err_msg, "error")
                    app.logger.error(err_msg)
                    return redirect(url_for("page_administration_donnees"))
            db.commit()
            msg_succes = (
                f"{ens_importes_count} enseignants ont été importés avec succès. "
                "Les anciens enseignants (y compris fictifs) et toutes les attributions existantes ont été supprimés."
            )
            flash(msg_succes, "success")

    except InvalidFileException:
        flash("Le fichier Excel des enseignants fourni est invalide ou corrompu.", "error")
    except AssertionError as ae:
        flash(f"Erreur de format ou de structure dans le fichier Excel des enseignants: {ae}", "error")
        app.logger.error(f"Erreur d'assertion lors de l'importation des enseignants: {ae}")
    except Exception as e:  # pylint: disable=broad-except
        if db and not db.closed and not db.autocommit:
            db.rollback()
        app.logger.error(f"Erreur inattendue lors de l'importation des enseignants: {type(e).__name__} - {e}")
        msg_err_inatt = f"Une erreur inattendue est survenue lors de l'importation des enseignants: {type(e).__name__}. L'opération a été annulée."
        flash(msg_err_inatt, "error")
    return redirect(url_for("page_administration_donnees"))


# --- Démarrage de l'application ---
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8080))
    debug_mode = os.environ.get("FLASK_DEBUG", "False").lower() == "true"
    app.run(host="0.0.0.0", port=port, debug=debug_mode)