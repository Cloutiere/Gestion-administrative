# mon_application/exports.py
"""
Ce module contient les fonctions de génération de fichiers Excel pour l'exportation.

Il est conçu pour être indépendant de Flask et se concentre uniquement sur la
création de documents Excel formatés à partir de données brutes fournies
en argument.
"""

import io
from typing import Any, cast

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def _apply_border_to_range(sheet: Worksheet, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    """Applique une bordure fine autour d'une plage de cellules."""
    thin_border_side = Side(style="thin")
    box_border = Border(
        left=thin_border_side,
        right=thin_border_side,
        top=thin_border_side,
        bottom=thin_border_side,
    )

    for row_iter in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row_iter:
            cell.border = box_border


def generer_export_taches(attributions_par_champ: dict[str, dict[str, Any]]) -> io.BytesIO:
    """
    Génère un fichier Excel des tâches attribuées, avec une feuille par champ.

    Args:
        attributions_par_champ: Dictionnaire des attributions groupées par champ.

    Returns:
        Un objet io.BytesIO contenant le fichier Excel (.xlsx) en mémoire.
    """
    workbook = openpyxl.Workbook()
    workbook.remove(cast(Worksheet, workbook.active))

    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    header_fill = PatternFill("solid", fgColor="4F81BD")
    header_align = Alignment(horizontal="center", vertical="center")

    cell_font = Font(name="Calibri", size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center")
    number_format_periods = "General"

    subtotal_font = Font(bold=True, name="Calibri", size=11)
    subtotal_fill = PatternFill("solid", fgColor="F2F2F2")

    grand_total_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    grand_total_fill = PatternFill("solid", fgColor="365F91")

    for champ_no, champ_data in attributions_par_champ.items():
        champ_nom = champ_data["nom"]
        attributions = champ_data["attributions"]
        nom_complet_champ = f"{champ_no}-{champ_nom}"
        safe_sheet_title = "".join(c for c in nom_complet_champ if c.isalnum() or c in " -_").strip()[:31]
        sheet: Worksheet = workbook.create_sheet(title=safe_sheet_title)

        current_row_num = 2
        headers = [
            "Enseignant",
            "Code cours",
            "Description",
            "Cours autre",
            "Nb. grp.",
            "Pér./ groupe",
            "Pér. Total",
        ]
        for col_idx, header_text in enumerate(headers, start=2):
            cell = sheet.cell(row=current_row_num, column=col_idx, value=header_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        current_row_num += 1

        grand_total_periodes_champ = 0.0
        subtotal_periodes_enseignant = 0.0
        previous_teacher_name = None
        previous_teacher_fullname = None
        group_start_row = current_row_num

        for attr in attributions:
            nom_enseignant_cle = f"{attr['nom']}, {attr['prenom']}"
            nom_enseignant_affichage = f"{attr['prenom']} {attr['nom']}"

            if previous_teacher_name is not None and nom_enseignant_cle != previous_teacher_name:
                sheet.cell(
                    row=current_row_num,
                    column=2,
                    value=f"Total pour {previous_teacher_fullname} ",
                )
                cell_value = sheet.cell(row=current_row_num, column=8)
                cell_value.value = subtotal_periodes_enseignant
                sheet.merge_cells(
                    start_row=current_row_num,
                    start_column=2,
                    end_row=current_row_num,
                    end_column=7,
                )
                for col_idx in range(2, 9):
                    cell = sheet.cell(row=current_row_num, column=col_idx)
                    cell.font = subtotal_font
                    cell.fill = subtotal_fill
                    if col_idx == 2:
                        cell.alignment = right_align
                    elif col_idx == 8:
                        cell.alignment = center_align
                        cell.number_format = number_format_periods
                _apply_border_to_range(sheet, group_start_row, current_row_num, 2, 8)
                current_row_num += 2
                group_start_row = current_row_num
                subtotal_periodes_enseignant = 0.0

            est_autre = "Oui" if attr["estcoursautre"] else "Non"
            nb_groupes = attr["total_groupes_pris"]
            per_groupe = float(attr["nbperiodes"])
            per_total_ligne = int(nb_groupes) * per_groupe

            row_data: list[Any] = [
                f"{attr['nom']}, {attr['prenom']}",
                attr["codecours"],
                attr["coursdescriptif"],
                est_autre,
                nb_groupes,
                per_groupe,
                per_total_ligne,
            ]
            for col_idx, cell_value in enumerate(row_data, start=2):
                cell = sheet.cell(row=current_row_num, column=col_idx, value=cell_value)
                cell.font = cell_font
                if col_idx in {2, 3, 4}:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align
                if col_idx in {6, 7, 8}:
                    cell.number_format = number_format_periods
            current_row_num += 1

            subtotal_periodes_enseignant += per_total_ligne
            grand_total_periodes_champ += per_total_ligne
            previous_teacher_name = nom_enseignant_cle
            previous_teacher_fullname = nom_enseignant_affichage

        if previous_teacher_name is not None:
            sheet.cell(
                row=current_row_num,
                column=2,
                value=f"Total pour {previous_teacher_fullname} ",
            )
            cell_value = sheet.cell(row=current_row_num, column=8)
            cell_value.value = subtotal_periodes_enseignant

            sheet.merge_cells(
                start_row=current_row_num,
                start_column=2,
                end_row=current_row_num,
                end_column=7,
            )
            for col_idx in range(2, 9):
                cell = sheet.cell(row=current_row_num, column=col_idx)
                cell.font = subtotal_font
                cell.fill = subtotal_fill
                if col_idx == 2:
                    cell.alignment = right_align
                elif col_idx == 8:
                    cell.alignment = center_align
                    cell.number_format = number_format_periods
            _apply_border_to_range(sheet, group_start_row, current_row_num, 2, 8)
            current_row_num += 2

            sheet.cell(
                row=current_row_num,
                column=2,
                value="TOTAL DES PÉRIODES ATTRIBUÉES DU CHAMP",
            )
            cell_grand_total_value = sheet.cell(row=current_row_num, column=8)
            cell_grand_total_value.value = grand_total_periodes_champ

            sheet.merge_cells(
                start_row=current_row_num,
                start_column=2,
                end_row=current_row_num,
                end_column=7,
            )
            for col_idx in range(2, 9):
                cell = sheet.cell(row=current_row_num, column=col_idx)
                cell.font = grand_total_font
                cell.fill = grand_total_fill
                if col_idx == 2:
                    cell.alignment = right_align
                elif col_idx == 8:
                    cell.alignment = center_align
                    cell.number_format = number_format_periods

        column_widths = {
            "A": 3,
            "B": 30,
            "C": 11,
            "D": 43,
            "E": 12,
            "F": 10,
            "G": 12,
            "H": 12,
        }
        for col_letter, width in column_widths.items():
            sheet.column_dimensions[col_letter].width = width
        sheet.freeze_panes = "B3"

    mem_file = io.BytesIO()
    workbook.save(mem_file)
    mem_file.seek(0)
    return mem_file


def generer_export_periodes_restantes(periodes_par_champ: dict[str, dict[str, Any]]) -> io.BytesIO:
    """
    Génère un fichier Excel des périodes restantes avec totaux et mise en forme.

    Args:
        periodes_par_champ: Dictionnaire des périodes restantes par champ.

    Returns:
        Un objet io.BytesIO contenant le fichier Excel (.xlsx) en mémoire.
    """
    workbook = openpyxl.Workbook()
    workbook.remove(cast(Worksheet, workbook.active))

    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    header_fill = PatternFill("solid", fgColor="4F81BD")
    header_align = Alignment(horizontal="center", vertical="center")
    cell_font = Font(name="Calibri", size=11)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    number_format_periods = "General"

    subtotal_font = Font(bold=True, name="Calibri", size=11)
    subtotal_fill = PatternFill("solid", fgColor="F2F2F2")

    grand_total_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    grand_total_fill = PatternFill("solid", fgColor="365F91")

    for champ_no, champ_data in periodes_par_champ.items():
        champ_nom = champ_data["nom"]
        periodes = champ_data["periodes"]
        nom_complet_champ = f"{champ_no}-{champ_nom}"
        safe_sheet_title = "".join(c for c in nom_complet_champ if c.isalnum() or c in " -_").strip()[:31]
        sheet: Worksheet = workbook.create_sheet(title=safe_sheet_title)

        current_row_num = 2
        headers = [
            "Champ",
            "Tâche restantes",
            "Code cours",
            "Description",
            "Cours autre",
            "Pér./ groupe",
        ]
        for col_idx, header_text in enumerate(headers, start=2):
            cell = sheet.cell(row=current_row_num, column=col_idx, value=header_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        current_row_num += 1

        grand_total_periodes = 0.0
        subtotal_periodes = 0.0
        previous_tache_raw = None
        previous_tache_display = None
        group_start_row = current_row_num

        for periode in periodes:
            current_tache_raw = periode["tache_restante"]
            prefix_champ = f"{champ_no}-"
            current_tache_display = current_tache_raw
            if current_tache_raw.startswith(prefix_champ):
                current_tache_display = current_tache_raw.removeprefix(prefix_champ)

            if previous_tache_raw is not None and current_tache_raw != previous_tache_raw:
                sheet.cell(
                    row=current_row_num,
                    column=2,
                    value=f"Total pour {previous_tache_display} ",
                )
                cell_value = sheet.cell(row=current_row_num, column=7)
                cell_value.value = subtotal_periodes
                sheet.merge_cells(
                    start_row=current_row_num,
                    start_column=2,
                    end_row=current_row_num,
                    end_column=6,
                )
                for col_idx in range(2, 8):
                    cell = sheet.cell(row=current_row_num, column=col_idx)
                    cell.font = subtotal_font
                    cell.fill = subtotal_fill
                    if col_idx == 2:
                        cell.alignment = right_align
                    elif col_idx == 7:
                        cell.alignment = center_align
                        cell.number_format = number_format_periods
                _apply_border_to_range(sheet, group_start_row, current_row_num, 2, 7)
                current_row_num += 2
                group_start_row = current_row_num
                subtotal_periodes = 0.0

            est_autre = "Oui" if periode["estcoursautre"] else "Non"
            current_periods_val = float(periode["nbperiodes"])
            row_data: list[Any] = [
                nom_complet_champ,
                current_tache_display,
                periode["codecours"],
                periode["coursdescriptif"],
                est_autre,
                current_periods_val,
            ]
            for col_idx, cell_value in enumerate(row_data, start=2):
                cell = sheet.cell(row=current_row_num, column=col_idx, value=cell_value)
                cell.font = cell_font
                if col_idx in {2, 3, 4, 5}:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align
                if col_idx == 7:
                    cell.number_format = number_format_periods
            current_row_num += 1

            subtotal_periodes += current_periods_val
            grand_total_periodes += current_periods_val
            previous_tache_raw = current_tache_raw
            previous_tache_display = current_tache_display

        if previous_tache_raw is not None:
            sheet.cell(
                row=current_row_num,
                column=2,
                value=f"Total pour {previous_tache_display} ",
            )
            cell_value = sheet.cell(row=current_row_num, column=7)
            cell_value.value = subtotal_periodes
            sheet.merge_cells(
                start_row=current_row_num,
                start_column=2,
                end_row=current_row_num,
                end_column=6,
            )
            for col_idx in range(2, 8):
                cell = sheet.cell(row=current_row_num, column=col_idx)
                cell.font = subtotal_font
                cell.fill = subtotal_fill
                if col_idx == 2:
                    cell.alignment = right_align
                elif col_idx == 7:
                    cell.alignment = center_align
                    cell.number_format = number_format_periods
            _apply_border_to_range(sheet, group_start_row, current_row_num, 2, 7)
            current_row_num += 2

            sheet.cell(
                row=current_row_num,
                column=2,
                value="TOTAL DES PÉRIODES RESTANTES DU CHAMP",
            )
            cell_grand_total_value = sheet.cell(row=current_row_num, column=7)
            cell_grand_total_value.value = grand_total_periodes
            sheet.merge_cells(
                start_row=current_row_num,
                start_column=2,
                end_row=current_row_num,
                end_column=6,
            )
            for col_idx in range(2, 8):
                cell = sheet.cell(row=current_row_num, column=col_idx)
                cell.font = grand_total_font
                cell.fill = grand_total_fill
                if col_idx == 2:
                    cell.alignment = right_align
                elif col_idx == 7:
                    cell.alignment = center_align
                    cell.number_format = number_format_periods

        column_widths = {
            "A": 3,
            "B": 35,
            "C": 16,
            "D": 11,
            "E": 43,
            "F": 12,
            "G": 12,
        }
        for col_letter, width in column_widths.items():
            sheet.column_dimensions[col_letter].width = width
        sheet.freeze_panes = "B3"

    mem_file = io.BytesIO()
    workbook.save(mem_file)
    mem_file.seek(0)
    return mem_file


def generer_export_org_scolaire(donnees_par_champ: dict[str, dict[str, Any]]) -> io.BytesIO:
    """
    Génère un fichier Excel pour l'organisation scolaire.

    Args:
        donnees_par_champ: Dictionnaire des données pivotées groupées par champ.

    Returns:
        Un objet io.BytesIO contenant le fichier Excel (.xlsx) en mémoire.
    """
    workbook = openpyxl.Workbook()
    workbook.remove(cast(Worksheet, workbook.active))

    total_font = Font(bold=True, name="Calibri", size=11)
    header_font_org = Font(bold=True, name="Calibri", size=11)
    header_align = Alignment(wrap_text=True, horizontal="center", vertical="center")

    HEADERS = [
        "NOM, PRÉNOM",
        "PÉRIODES RÉGULIER",
        "PÉRIODES ADAPTATION SCOLAIRE",
        "PÉRIODES SPORT-ÉTUDES",
        "PÉRIODES ENSEIGNANT RESSOURCE",
        "PÉRIODES AIDESEC",
        "PÉRIODES DIPLÔMA",
        "PÉRIODES MESURE SEUIL (UTILISÉE COORDINATION PP)",
        "PÉRIODES MESURE SEUIL (RESSOURCES AUTRES)",
        "PÉRIODES MESURE SEUIL (POUR FABLAB)",
        "PÉRIODES MESURE SEUIL (BONIFIER ALTERNE)",
        "PÉRIODES ALTERNE",
        "PÉRIODES FORMANUM",
        "PÉRIODES MENTORAT",
        "PÉRIODES COORDINATION SPORT-ÉTUDES",
        "PÉRIODES SOUTIEN SPORT-ÉTUDES",
    ]
    COLUMN_WIDTH = 22

    for champ_no, champ_data in donnees_par_champ.items():
        champ_nom = champ_data["nom"]
        donnees = champ_data["donnees"]
        nom_complet_champ = f"{champ_no}-{champ_nom}"
        safe_sheet_title = "".join(c for c in nom_complet_champ if c.isalnum() or c in " -_").strip()[:31]
        sheet: Worksheet = workbook.create_sheet(title=safe_sheet_title)

        for col_idx, header_text in enumerate(HEADERS, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header_text)
            cell.font = header_font_org
            cell.alignment = header_align

        totals: dict[str, float] = {h: 0.0 for h in HEADERS if h != "NOM, PRÉNOM"}

        for item in donnees:
            row_data: list[str | float] = []
            if item["estfictif"]:
                row_data.append(item["nomcomplet"])
            else:
                row_data.append(f"{item['nom']}, {item['prenom']}")

            for header in HEADERS[1:]:
                period_value = item.get(header, 0.0)
                row_data.append(period_value)
                totals[header] += period_value
            sheet.append(row_data)

        if donnees:
            total_row_idx = sheet.max_row + 1
            total_cell = sheet.cell(row=total_row_idx, column=1, value="TOTAL")
            total_cell.font = total_font
            for col_idx, header_text in enumerate(HEADERS, start=1):
                if header_text in totals:
                    cell = sheet.cell(row=total_row_idx, column=col_idx)
                    cell.value = totals[header_text]
                    cell.font = total_font

            grand_total_row_idx = total_row_idx + 2
            grand_total_periodes_champ = sum(totals.values())
            label_cell = sheet.cell(
                row=grand_total_row_idx,
                column=1,
                value="TOTAL PÉRIODES DU CHAMP",
            )
            label_cell.font = total_font
            value_cell = sheet.cell(
                row=grand_total_row_idx,
                column=2,
            )
            value_cell.value = grand_total_periodes_champ
            value_cell.font = total_font

        for i in range(1, len(HEADERS) + 1):
            letter = get_column_letter(i)
            sheet.column_dimensions[letter].width = COLUMN_WIDTH

    mem_file = io.BytesIO()
    workbook.save(mem_file)
    mem_file.seek(0)
    return mem_file
