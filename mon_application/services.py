# mon_application/services.py
"""
Ce module contient la logique métier de l'application (couche de services).

Il a pour but de découpler la logique complexe des routes Flask (contrôleurs).
Les fonctions ici gèrent des tâches comme le traitement de fichiers, les
opérations de base de données transactionnelles complexes, l'application des
règles de gestion et l'orchestration des appels à la couche d'accès aux
données (DAO).
"""

from collections import defaultdict
from typing import Any, cast

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import case, func
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import aliased, joinedload

from .extensions import db

# NOUVEAU: Import des modèles requis pour la refactorisation
from .models import (
    AnneeScolaire,
    AttributionCours,
    Champ,
    ChampAnneeStatut,
    Cours,
    Enseignant,
    PreparationHoraire,
    TypeFinancement,
    User,
)


# --- Exceptions Personnalisées pour la Couche de Service ---
class ServiceException(Exception):
    """Exception de base pour les erreurs de la couche de service."""

    def __init__(self, message="Une erreur est survenue."):
        self.message = message
        super().__init__(self.message)


class EntityNotFoundError(ServiceException):
    """Levée lorsqu'une entité n'est pas trouvée."""

    def __init__(self, message="L'entité n'a pas été trouvée."):
        super().__init__(message)


class DuplicateEntityError(ServiceException):
    """Levée lors d'une tentative de création d'une entité qui existe déjà."""

    def __init__(self, message="Cette entité existe déjà."):
        super().__init__(message)


class BusinessRuleValidationError(ServiceException):
    """Levée lorsqu'une règle métier est violée."""

    def __init__(self, message="Opération non autorisée par les règles de gestion."):
        super().__init__(message)


class ForeignKeyError(ServiceException):
    """Levée lorsqu'une suppression est bloquée par une contrainte de clé étrangère."""

    def __init__(self, message="Impossible de supprimer, cette entité est en cours d'utilisation."):
        self.message = message
        super().__init__(self.message)


# --- Services - Utilisateurs et Rôles ---
def register_first_admin_service(username: str, password: str, confirm_password: str) -> User:
    """
    Gère la logique d'inscription du premier admin.
    Lève des exceptions et retourne l'objet User prêt à être commit.
    """
    if db.session.query(User.id).count() > 0:
        raise BusinessRuleValidationError("L'inscription n'est autorisée que pour le premier utilisateur.")

    if not all([username, password, confirm_password]):
        raise BusinessRuleValidationError("Tous les champs sont requis.")
    if password != confirm_password:
        raise BusinessRuleValidationError("Les mots de passe ne correspondent pas.")
    if len(password) < 6:
        raise BusinessRuleValidationError("Le mot de passe doit contenir au moins 6 caractères.")

    if db.session.query(User).filter_by(username=username).first():
        raise DuplicateEntityError("Ce nom d'utilisateur est déjà pris.")

    new_admin = User(username=username, is_admin=True)
    new_admin.set_password(password)

    db.session.add(new_admin)
    return new_admin


# --- Services - Traitement des fichiers et Importations ---
def process_courses_excel(file_stream: Any) -> list[dict[str, Any]]:
    nouveaux_cours: list[dict[str, Any]] = []
    try:
        workbook = openpyxl.load_workbook(file_stream)
        sheet = cast(Worksheet, workbook.active)
        if sheet.max_row <= 1:
            raise ValueError("Fichier Excel vide ou ne contenant que l'en-tête.")
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            values = [cell.value for cell in row]
            if not any(v is not None and str(v).strip() != "" for v in values[:7]):
                continue
            (
                champ_no_raw,
                code_cours_raw,
                desc_raw,
                nb_grp_raw,
                nb_per_raw,
            ) = (values[0], values[1], values[3], values[4], values[5])
            est_autre_raw = values[6] if len(values) > 6 else None
            financement_code_raw = values[7] if len(values) > 7 else None
            if not all([champ_no_raw, code_cours_raw, desc_raw, nb_grp_raw, nb_per_raw]):
                raise ValueError(f"Ligne {row_idx}: Données essentielles manquantes.")
            try:
                est_autre = str(est_autre_raw).strip().upper() in ("VRAI", "TRUE", "OUI", "YES", "1")
                financement_code = str(financement_code_raw).strip() if financement_code_raw else None
                nouveaux_cours.append(
                    {
                        "codecours": str(code_cours_raw).strip(),
                        "champno": str(champ_no_raw).strip(),
                        "coursdescriptif": str(desc_raw).strip(),
                        "nbperiodes": float(str(nb_per_raw).replace(",", ".")),
                        "nbgroupeinitial": int(float(str(nb_grp_raw).replace(",", "."))),
                        "estcoursautre": est_autre,
                        "financement_code": financement_code,
                    }
                )
            except (ValueError, TypeError) as e:
                raise ValueError(f"Ligne {row_idx}: Erreur de type de données. Détails: {e}") from e
    except InvalidFileException as e:
        raise InvalidFileException("Fichier Excel corrompu ou invalide.") from e
    if not nouveaux_cours:
        raise ValueError("Aucun cours valide n'a été trouvé dans le fichier.")
    return nouveaux_cours


class ImportationStats:
    def __init__(self) -> None:
        self.imported_count = 0
        self.deleted_attributions_count = 0
        self.deleted_main_entities_count = 0


def save_imported_courses(courses_data: list[dict[str, Any]], annee_id: int) -> ImportationStats:
    """
    Sauvegarde les cours importés en utilisant l'ORM SQLAlchemy.
    Supprime les anciens cours et attributions de l'année avant d'insérer les nouveaux.
    L'opération est transactionnelle.
    """
    stats = ImportationStats()
    try:
        # Étape 1 : Compter les attributions qui seront supprimées par cascade pour les stats.
        # On le fait avant la suppression.
        stats.deleted_attributions_count = db.session.query(AttributionCours.attributionid).filter(AttributionCours.annee_id_cours == annee_id).count()

        # Étape 2 : Supprimer tous les cours de l'année.
        # L'option synchronize_session=False est plus performante pour les suppressions en masse.
        # La cascade configurée sur le modèle AnneeScolaire -> Cours s'occupera de tout.
        # Mais ici, on cible directement les cours pour plus de clarté.
        deleted_rows = db.session.query(Cours).filter_by(annee_id=annee_id).delete(synchronize_session=False)
        stats.deleted_main_entities_count = deleted_rows

        # Étape 3 : Créer et ajouter les nouveaux cours
        nouveaux_cours_objets = [Cours(annee_id=annee_id, **data) for data in courses_data]
        if nouveaux_cours_objets:
            db.session.add_all(nouveaux_cours_objets)
            stats.imported_count = len(nouveaux_cours_objets)

        # Étape 4 : Valider la transaction
        db.session.commit()
        return stats
    except IntegrityError as e:
        db.session.rollback()
        # Cela peut se produire si un `financement_code` ou `champno` n'existe pas.
        raise ServiceException(f"Erreur d'intégrité des données: un champ ou un code de financement est-il invalide ? Détails: {e}")
    except Exception as e:
        db.session.rollback()
        raise ServiceException(f"Erreur de base de données lors de l'importation des cours: {e}")


def process_teachers_excel(file_stream: Any) -> list[dict[str, Any]]:
    nouveaux_enseignants: list[dict[str, Any]] = []
    try:
        workbook = openpyxl.load_workbook(file_stream)
        sheet = cast(Worksheet, workbook.active)
        if sheet.max_row <= 1:
            raise ValueError("Fichier Excel vide ou ne contenant que l'en-tête.")
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            values = [cell.value for cell in row]
            if not any(v is not None and str(v).strip() != "" for v in values[:4]):
                continue
            champ_no_raw, nom_raw, prenom_raw, temps_plein_raw = (
                values[0],
                values[1],
                values[2],
                values[3],
            )
            if not all([champ_no_raw, nom_raw, prenom_raw, temps_plein_raw is not None]):
                raise ValueError(f"Ligne {row_idx}: Données essentielles manquantes.")
            try:
                nom_clean, prenom_clean = str(nom_raw).strip(), str(prenom_raw).strip()
                if not nom_clean or not prenom_clean:
                    continue
                nouveaux_enseignants.append(
                    {
                        "nom": nom_clean,
                        "prenom": prenom_clean,
                        "champno": str(champ_no_raw).strip(),
                        "esttempsplein": str(temps_plein_raw).strip().upper() in ("VRAI", "TRUE", "OUI", "YES", "1"),
                    }
                )
            except (ValueError, TypeError) as e:
                raise ValueError(f"Ligne {row_idx}: Erreur de type de données. Détails: {e}") from e
    except InvalidFileException as e:
        raise InvalidFileException("Fichier Excel des enseignants corrompu ou invalide.") from e
    if not nouveaux_enseignants:
        raise ValueError("Aucun enseignant valide n'a été trouvé dans le fichier.")
    return nouveaux_enseignants


def save_imported_teachers(teachers_data: list[dict[str, Any]], annee_id: int) -> ImportationStats:
    """
    Sauvegarde les enseignants importés en utilisant l'ORM SQLAlchemy.
    Supprime les anciens enseignants et attributions de l'année avant d'insérer les nouveaux.
    L'opération est transactionnelle.
    """
    stats = ImportationStats()
    try:
        # Étape 1 : Compter les attributions qui seront supprimées par cascade
        stats.deleted_attributions_count = (
            db.session.query(AttributionCours.attributionid).join(Enseignant, Enseignant.enseignantid == AttributionCours.enseignantid).filter(Enseignant.annee_id == annee_id).count()
        )

        # Étape 2 : Supprimer tous les enseignants de l'année.
        deleted_rows = db.session.query(Enseignant).filter_by(annee_id=annee_id).delete(synchronize_session=False)
        stats.deleted_main_entities_count = deleted_rows

        # Étape 3 : Créer et ajouter les nouveaux enseignants
        nouveaux_enseignants_objets = []
        for data in teachers_data:
            data["nomcomplet"] = f"{data['prenom']} {data['nom']}"
            nouveaux_enseignants_objets.append(Enseignant(annee_id=annee_id, **data))

        if nouveaux_enseignants_objets:
            db.session.add_all(nouveaux_enseignants_objets)
            stats.imported_count = len(nouveaux_enseignants_objets)

        # Étape 4 : Valider la transaction
        db.session.commit()
        return stats
    except IntegrityError as e:
        db.session.rollback()
        # Cause probable : `champno` invalide ou doublon nom/prénom pour la même année.
        raise ServiceException(f"Erreur d'intégrité: un champ est-il invalide ou y a-t-il un doublon de nom/prénom? Détails: {e}")
    except Exception as e:
        db.session.rollback()
        raise ServiceException(f"Erreur de base de données lors de l'importation des enseignants: {e}")


# --- SECTION REFACTORISÉE : Années Scolaires avec ORM ---


def get_all_annees_service() -> list[dict[str, Any]]:
    """Récupère toutes les années scolaires via l'ORM, ordonnées par libellé décroissant."""
    try:
        annees = db.session.query(AnneeScolaire).order_by(AnneeScolaire.libelle_annee.desc()).all()
        return [{"annee_id": a.annee_id, "libelle_annee": a.libelle_annee, "est_courante": a.est_courante} for a in annees]
    except Exception as e:
        raise ServiceException(f"Erreur ORM lors de la récupération des années: {e}")


def get_active_year_service() -> dict[str, Any]:
    """Récupère l'année scolaire active via l'ORM."""
    try:
        active_year = db.session.query(AnneeScolaire).filter_by(est_courante=True).one_or_none()
        if not active_year:
            raise BusinessRuleValidationError("Aucune année scolaire active n'est définie. Veuillez en configurer une dans le panneau d'administration.")
        return {
            "annee_id": active_year.annee_id,
            "libelle_annee": active_year.libelle_annee,
            "est_courante": active_year.est_courante,
        }
    except Exception as e:
        # Re-lever les exceptions métier, sinon encapsuler
        if isinstance(e, BusinessRuleValidationError):
            raise
        raise ServiceException(f"Erreur ORM lors de la récupération de l'année active : {e}")


def determine_active_school_year_service(all_years: list[dict[str, Any]], has_dashboard_access: bool, annee_id_session: int | None) -> tuple[dict[str, Any] | None, str | None]:
    """
    Détermine l'année scolaire active à afficher en se basant sur la logique métier.
    Retourne l'année active (dict) et un message d'avertissement potentiel (str).
    """
    annee_active = None
    warning_message = None
    annee_courante_existante = False

    # 1. Tenter de trouver l'année de la session si l'utilisateur y a droit
    if has_dashboard_access and annee_id_session:
        annee_active = next((annee for annee in all_years if annee["annee_id"] == annee_id_session), None)

    # 2. Si non trouvée, chercher l'année marquée comme "courante"
    if not annee_active:
        annee_courante = next((annee for annee in all_years if annee["est_courante"]), None)
        if annee_courante:
            annee_active = annee_courante
            annee_courante_existante = True

    # 3. Si toujours pas d'année, prendre la plus récente comme solution de repli
    if not annee_active and all_years:
        # La liste est déjà triée par libellé décroissant par `get_all_annees_service`
        annee_active = all_years[0]
        if has_dashboard_access and not annee_courante_existante:
            warning_message = "Aucune année scolaire n'est définie comme 'courante'. Affichage de la plus récente par défaut."

    return annee_active, warning_message


def create_annee_scolaire_service(libelle: str) -> dict[str, Any]:
    """
    Crée une nouvelle année scolaire via l'ORM.
    Si aucune année n'est courante, la nouvelle année le devient automatiquement.
    """
    if db.session.query(AnneeScolaire).filter_by(libelle_annee=libelle).first():
        raise DuplicateEntityError(f"L'année '{libelle}' existe déjà.")

    try:
        annee_courante_existante = db.session.query(AnneeScolaire).filter_by(est_courante=True).first()

        new_annee = AnneeScolaire(libelle_annee=libelle)
        if not annee_courante_existante:
            new_annee.est_courante = True

        db.session.add(new_annee)
        db.session.commit()

        return {
            "annee_id": new_annee.annee_id,
            "libelle_annee": new_annee.libelle_annee,
            "est_courante": new_annee.est_courante,
        }
    except IntegrityError:
        db.session.rollback()
        raise DuplicateEntityError(f"L'année '{libelle}' existe déjà.")
    except Exception as e:
        db.session.rollback()
        raise ServiceException(f"Erreur de base de données lors de la création de l'année: {e}")


def set_annee_courante_service(annee_id: int) -> None:
    """
    Définit une année scolaire comme courante via l'ORM, en assurant une transaction atomique.
    """
    annee_a_definir = db.session.get(AnneeScolaire, annee_id)
    if not annee_a_definir:
        raise EntityNotFoundError("Année non trouvée.")

    ancienne_annee_courante = db.session.query(AnneeScolaire).filter_by(est_courante=True).first()

    if ancienne_annee_courante and ancienne_annee_courante.annee_id != annee_a_definir.annee_id:
        ancienne_annee_courante.est_courante = False

    annee_a_definir.est_courante = True

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        raise ServiceException(f"La mise à jour de l'année courante a échoué en base de données : {e}")


def get_all_champs_service() -> list[dict[str, Any]]:
    try:
        champs_orm = db.session.query(Champ).order_by(Champ.champno).all()
        return [{"champno": champ.champno, "champnom": champ.champnom} for champ in champs_orm]
    except Exception as e:
        raise ServiceException(f"Erreur ORM lors de la récupération des champs : {e}")


def get_champ_details_service(champ_no: str, annee_id: int) -> dict[str, Any]:
    """Récupère les détails d'un champ et ses statuts pour une année donnée via l'ORM."""
    champ = db.session.get(Champ, champ_no)
    if not champ:
        raise EntityNotFoundError(f"Le champ '{champ_no}' n'a pas été trouvé.")

    statut = db.session.query(ChampAnneeStatut).filter_by(champ_no=champ_no, annee_id=annee_id).first()

    return {
        "ChampNo": champ.champno,
        "ChampNom": champ.champnom,
        "est_verrouille": statut.est_verrouille if statut else False,
        "est_confirme": statut.est_confirme if statut else False,
    }


def get_all_champ_statuses_for_year_service(annee_id: int) -> dict[str, dict[str, bool]]:
    """Récupère les statuts (verrouillé/confirmé) de tous les champs pour une année via l'ORM."""
    try:
        statuses_orm = db.session.query(ChampAnneeStatut).filter_by(annee_id=annee_id).all()
        return {status.champ_no: {"est_verrouille": status.est_verrouille, "est_confirme": status.est_confirme} for status in statuses_orm}
    except Exception as e:
        raise ServiceException(f"Erreur ORM lors de la récupération des statuts de champs : {e}")


def _toggle_champ_status_service(champ_no: str, annee_id: int, status_attribute: str) -> bool:
    """
    Fonction helper pour basculer un statut booléen sur ChampAnneeStatut.
    Crée l'entrée si elle n'existe pas (logique UPSERT).
    """
    if status_attribute not in ("est_verrouille", "est_confirme"):
        raise ValueError("Attribut de statut invalide.")

    status = db.session.query(ChampAnneeStatut).filter_by(champ_no=champ_no, annee_id=annee_id).first()

    new_value: bool
    if status:
        current_value = getattr(status, status_attribute)
        new_value = not current_value
        setattr(status, status_attribute, new_value)
    else:
        # Le premier toggle met toujours à True
        new_value = True
        status = ChampAnneeStatut(champ_no=champ_no, annee_id=annee_id)
        setattr(status, status_attribute, new_value)
        db.session.add(status)

    try:
        db.session.commit()
        return new_value
    except IntegrityError as e:
        db.session.rollback()
        raise ServiceException(f"Erreur d'intégrité : le champ ou l'année n'existe pas. Détails: {e}")
    except Exception as e:
        db.session.rollback()
        raise ServiceException(f"Erreur de base de données lors de la mise à jour du statut: {e}")


def toggle_champ_lock_service(champ_no: str, annee_id: int) -> bool:
    """Bascule le statut de verrouillage d'un champ pour une année donnée via l'ORM."""
    try:
        return _toggle_champ_status_service(champ_no, annee_id, "est_verrouille")
    except ValueError:  # Devrait être impossible, mais par sécurité
        raise ServiceException("Erreur interne: attribut de statut invalide.")


def toggle_champ_confirm_service(champ_no: str, annee_id: int) -> bool:
    """Bascule le statut de confirmation d'un champ pour une année donnée via l'ORM."""
    try:
        return _toggle_champ_status_service(champ_no, annee_id, "est_confirme")
    except ValueError:
        raise ServiceException("Erreur interne: attribut de statut invalide.")


def _cours_to_dict(cours: Cours) -> dict[str, Any]:
    """Utilitaire pour convertir un objet Cours en dictionnaire."""
    return {
        "codecours": cours.codecours,
        "annee_id": cours.annee_id,
        "champno": cours.champno,
        "coursdescriptif": cours.coursdescriptif,
        "nbperiodes": float(cours.nbperiodes),
        "nbgroupeinitial": cours.nbgroupeinitial,
        "estcoursautre": cours.estcoursautre,
        "financement_code": cours.financement_code,
    }


def get_course_details_service(code_cours: str, annee_id: int) -> dict[str, Any]:
    """Récupère les détails d'un cours via l'ORM."""
    cours = db.session.get(Cours, {"codecours": code_cours, "annee_id": annee_id})
    if not cours:
        raise EntityNotFoundError("Cours non trouvé pour cette année.")
    return _cours_to_dict(cours)


def create_course_service(data: dict[str, Any], annee_id: int) -> dict[str, Any]:
    """Crée un nouveau cours via l'ORM."""
    try:
        new_cours = Cours(
            codecours=data["codecours"],
            annee_id=annee_id,
            champno=data["champno"],
            coursdescriptif=data["coursdescriptif"],
            nbperiodes=data["nbperiodes"],
            nbgroupeinitial=data["nbgroupeinitial"],
            estcoursautre=data["estcoursautre"],
            financement_code=data.get("financement_code") or None,
        )
        db.session.add(new_cours)
        db.session.commit()
        return _cours_to_dict(new_cours)
    except IntegrityError:
        db.session.rollback()
        raise DuplicateEntityError("Un cours avec ce code existe déjà pour cette année.")
    except Exception as e:
        db.session.rollback()
        raise ServiceException(f"Erreur de base de données lors de la création du cours : {e}")


def update_course_service(code_cours: str, annee_id: int, data: dict[str, Any]) -> dict[str, Any]:
    """Met à jour un cours via l'ORM."""
    cours_to_update = db.session.get(Cours, {"codecours": code_cours, "annee_id": annee_id})
    if not cours_to_update:
        raise EntityNotFoundError("Cours non trouvé pour cette année.")

    try:
        cours_to_update.champno = data["champno"]
        cours_to_update.coursdescriptif = data["coursdescriptif"]
        cours_to_update.nbperiodes = data["nbperiodes"]
        cours_to_update.nbgroupeinitial = data["nbgroupeinitial"]
        cours_to_update.estcoursautre = data["estcoursautre"]
        cours_to_update.financement_code = data.get("financement_code") or None
        db.session.commit()
        return _cours_to_dict(cours_to_update)
    except Exception as e:
        db.session.rollback()
        raise ServiceException(f"Erreur de base de données lors de la mise à jour du cours : {e}")


def delete_course_service(code_cours: str, annee_id: int) -> None:
    """Supprime un cours via l'ORM."""
    cours_to_delete = db.session.get(Cours, {"codecours": code_cours, "annee_id": annee_id})
    if not cours_to_delete:
        raise EntityNotFoundError("Cours non trouvé pour cette année.")

    try:
        db.session.delete(cours_to_delete)
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        raise ForeignKeyError("Impossible de supprimer : ce cours est attribué à un ou plusieurs enseignants.")
    except Exception as e:
        db.session.rollback()
        raise ServiceException(f"Erreur de base de données lors de la suppression du cours : {e}")


def reassign_course_to_champ_service(code_cours: str, annee_id: int, nouveau_champ_no: str) -> dict[str, Any]:
    """Réassigne un cours à un nouveau champ en utilisant l'ORM."""
    cours = db.session.get(Cours, {"codecours": code_cours, "annee_id": annee_id})
    if not cours:
        raise EntityNotFoundError(f"Cours {code_cours} non trouvé pour l'année {annee_id}.")

    nouveau_champ = db.session.get(Champ, nouveau_champ_no)
    if not nouveau_champ:
        raise BusinessRuleValidationError(f"Le champ de destination '{nouveau_champ_no}' est invalide.")

    try:
        cours.champno = nouveau_champ_no
        db.session.commit()
        return {
            "nouveau_champ_no": nouveau_champ.champno,
            "nouveau_champ_nom": nouveau_champ.champnom,
        }
    except Exception as e:
        db.session.rollback()
        raise ServiceException(f"Erreur de base de données lors de la réassignation du cours: {e}")


def reassign_course_to_financement_service(code_cours: str, annee_id: int, code_financement: str | None) -> None:
    """Réassigne un cours à un nouveau type de financement en utilisant l'ORM."""
    cours = db.session.get(Cours, {"codecours": code_cours, "annee_id": annee_id})
    if not cours:
        raise EntityNotFoundError(f"Cours {code_cours} non trouvé pour l'année {annee_id}.")

    # Valider que le financement existe s'il n'est pas None
    if code_financement and not db.session.get(TypeFinancement, code_financement):
        raise BusinessRuleValidationError(f"Le type de financement '{code_financement}' est invalide.")

    try:
        cours.financement_code = code_financement
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        raise ServiceException(f"Erreur de base de données lors de la réassignation du financement: {e}")


def _enseignant_to_dict(enseignant: Enseignant) -> dict[str, Any]:
    """Utilitaire pour convertir un objet Enseignant en dictionnaire."""
    return {
        "enseignantid": enseignant.enseignantid,
        "annee_id": enseignant.annee_id,
        "nomcomplet": enseignant.nomcomplet,
        "nom": enseignant.nom,
        "prenom": enseignant.prenom,
        "champno": enseignant.champno,
        "esttempsplein": enseignant.esttempsplein,
        "estfictif": enseignant.estfictif,
    }


def get_teacher_details_service(enseignant_id: int) -> dict[str, Any]:
    """Récupère les détails d'un enseignant (non fictif) via l'ORM."""
    enseignant = db.session.query(Enseignant).filter_by(enseignantid=enseignant_id, estfictif=False).first()
    if not enseignant:
        raise EntityNotFoundError("Enseignant non trouvé ou non modifiable.")
    return _enseignant_to_dict(enseignant)


def get_any_teacher_by_id_service(enseignant_id: int) -> dict[str, Any]:
    """Récupère les détails de n'importe quel enseignant (réel ou fictif) par son ID."""
    enseignant = db.session.get(Enseignant, enseignant_id)
    if not enseignant:
        raise EntityNotFoundError("Enseignant non trouvé.")
    return _enseignant_to_dict(enseignant)


def create_teacher_service(data: dict[str, Any], annee_id: int) -> dict[str, Any]:
    """Crée un nouvel enseignant via l'ORM."""
    try:
        new_teacher = Enseignant(
            annee_id=annee_id,
            nom=data["nom"],
            prenom=data["prenom"],
            nomcomplet=f"{data['prenom']} {data['nom']}",
            champno=data["champno"],
            esttempsplein=data["esttempsplein"],
            estfictif=False,
        )
        db.session.add(new_teacher)
        db.session.commit()
        return _enseignant_to_dict(new_teacher)
    except IntegrityError:
        db.session.rollback()
        raise DuplicateEntityError("Un enseignant avec ce nom/prénom existe déjà pour cette année.")
    except Exception as e:
        db.session.rollback()
        raise ServiceException(f"Erreur de base de données lors de la création de l'enseignant : {e}")


def update_teacher_service(enseignant_id: int, data: dict[str, Any]) -> dict[str, Any]:
    """Met à jour un enseignant (non fictif) via l'ORM."""
    teacher_to_update = db.session.query(Enseignant).filter_by(enseignantid=enseignant_id, estfictif=False).first()
    if not teacher_to_update:
        raise EntityNotFoundError("Enseignant non trouvé ou non modifiable (fictif).")

    try:
        teacher_to_update.nom = data["nom"]
        teacher_to_update.prenom = data["prenom"]
        teacher_to_update.nomcomplet = f"{data['prenom']} {data['nom']}"
        teacher_to_update.champno = data["champno"]
        teacher_to_update.esttempsplein = data["esttempsplein"]
        db.session.commit()
        return _enseignant_to_dict(teacher_to_update)
    except IntegrityError:
        db.session.rollback()
        raise DuplicateEntityError("Un autre enseignant avec ce nom/prénom existe déjà pour cette année.")
    except Exception as e:
        db.session.rollback()
        raise ServiceException(f"Erreur de base de données lors de la mise à jour de l'enseignant : {e}")


def delete_teacher_service(enseignant_id: int) -> list[dict[str, Any]]:
    """Supprime un enseignant via l'ORM et retourne les cours qui lui étaient affectés."""
    teacher_to_delete = db.session.query(Enseignant).options(joinedload(Enseignant.attributions)).filter_by(enseignantid=enseignant_id).first()

    if not teacher_to_delete:
        raise EntityNotFoundError("Enseignant non trouvé.")

    # La règle métier est de retourner la liste des cours affectés avant suppression.
    cours_affectes = [{"codecours": attr.codecours, "annee_id_cours": attr.annee_id_cours} for attr in teacher_to_delete.attributions]

    try:
        db.session.delete(teacher_to_delete)
        db.session.commit()
        return cours_affectes
    except Exception as e:
        db.session.rollback()
        # Il est peu probable d'avoir une IntegrityError ici à cause de la cascade,
        # mais une gestion d'erreur générique est une bonne pratique.
        raise ServiceException(f"Erreur de base de données lors de la suppression: {e}")


def create_fictitious_teacher_service(champ_no: str, annee_id: int) -> dict[str, Any]:
    """Crée un enseignant fictif (tâche) pour un champ/année via l'ORM."""
    try:
        # 1. Récupérer les tâches existantes pour ce champ avec l'ORM
        fictifs_existants = (
            db.session.query(Enseignant)
            .filter(
                Enseignant.champno == champ_no,
                Enseignant.annee_id == annee_id,
                Enseignant.estfictif == True,  # noqa: E712
                Enseignant.nomcomplet.like(f"{champ_no}-Tâche restante-%"),
            )
            .all()
        )

        # 2. Déterminer le prochain numéro disponible
        numeros = [int(f.nomcomplet.split("-")[-1]) for f in fictifs_existants if f.nomcomplet.split("-")[-1].isdigit()]
        next_num = max(numeros) + 1 if numeros else 1
        nom_tache = f"{champ_no}-Tâche restante-{next_num}"

        # 3. Créer le nouvel objet Enseignant
        nouveau_fictif = Enseignant(
            annee_id=annee_id,
            nomcomplet=nom_tache,
            champno=champ_no,
            estfictif=True,
            esttempsplein=True,  # Les tâches sont considérées TP par défaut
        )

        # 4. Sauvegarder et commiter la transaction
        db.session.add(nouveau_fictif)
        db.session.commit()

        return _enseignant_to_dict(nouveau_fictif)
    except IntegrityError as e:
        db.session.rollback()
        # Devrait être rare, mais peut arriver si le champ_no n'existe pas
        raise ServiceException(f"Erreur d'intégrité lors de la création de la tâche: {e}")
    except Exception as e:
        db.session.rollback()
        raise ServiceException(f"Erreur de base de données lors de la création de la tâche: {e}")


def get_all_financements_service() -> list[dict[str, Any]]:
    """Récupère tous les types de financement via l'ORM."""
    try:
        financements = db.session.query(TypeFinancement).order_by(TypeFinancement.code).all()
        return [{"code": f.code, "libelle": f.libelle} for f in financements]
    except Exception as e:
        raise ServiceException(f"Erreur ORM lors de la récupération des financements: {e}")


def create_financement_service(code: str, libelle: str) -> dict[str, Any]:
    """Crée un nouveau type de financement via l'ORM."""
    new_financement = TypeFinancement(code=code, libelle=libelle)
    db.session.add(new_financement)
    try:
        db.session.commit()
        return {"code": new_financement.code, "libelle": new_financement.libelle}
    except IntegrityError:
        db.session.rollback()
        raise DuplicateEntityError("Ce code de financement existe déjà.")
    except Exception as e:
        db.session.rollback()
        raise ServiceException(f"Erreur de base de données: {e}")


def update_financement_service(code: str, libelle: str) -> dict[str, Any]:
    """Met à jour un type de financement via l'ORM."""
    financement = db.session.get(TypeFinancement, code)
    if not financement:
        raise EntityNotFoundError("Type de financement non trouvé.")

    financement.libelle = libelle
    try:
        db.session.commit()
        return {"code": financement.code, "libelle": financement.libelle}
    except Exception as e:
        db.session.rollback()
        raise ServiceException(f"Erreur de base de données: {e}")


def delete_financement_service(code: str) -> None:
    """Supprime un type de financement via l'ORM."""
    financement = db.session.get(TypeFinancement, code)
    if not financement:
        raise EntityNotFoundError("Type de financement non trouvé.")

    db.session.delete(financement)
    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        raise ForeignKeyError("Impossible de supprimer : ce financement est utilisé par des cours.")
    except Exception as e:
        db.session.rollback()
        raise ServiceException(f"Erreur de base de données: {e}")


def get_all_users_with_details_service() -> dict[str, Any]:
    """Récupère tous les utilisateurs avec des détails et le décompte des admins, via l'ORM."""
    try:
        users_orm = db.session.query(User).options(joinedload(User.champs_autorises)).order_by(User.username).all()
        users_data = [
            {
                "id": user.id,
                "username": user.username,
                "is_admin": user.is_admin,
                "is_dashboard_only": user.is_dashboard_only,
                "allowed_champs": [c.champno for c in user.champs_autorises],
            }
            for user in users_orm
        ]

        admin_count = db.session.query(User).filter_by(is_admin=True).count()

        return {"users": users_data, "admin_count": admin_count}
    except Exception as e:
        raise ServiceException(f"Erreur ORM lors de la récupération des détails utilisateurs : {e}")


def create_user_service(username: str, password: str, role: str, allowed_champs: list[str]) -> dict[str, Any]:
    """Crée un nouvel utilisateur en utilisant l'ORM SQLAlchemy."""
    if len(password) < 6:
        raise BusinessRuleValidationError("Le mot de passe doit faire au moins 6 caractères.")

    if db.session.query(User).filter_by(username=username).first():
        raise DuplicateEntityError("Ce nom d'utilisateur est déjà pris.")

    new_user = User(username=username)
    new_user.set_password(password)

    if role == "admin":
        new_user.is_admin = True
    elif role == "dashboard_only":
        new_user.is_dashboard_only = True
    elif role == "specific_champs" and allowed_champs:
        champs_objects = db.session.query(Champ).filter(Champ.champno.in_(allowed_champs)).all()
        if len(champs_objects) != len(allowed_champs):
            raise BusinessRuleValidationError("Un ou plusieurs champs spécifiés sont invalides.")
        new_user.champs_autorises = champs_objects

    try:
        db.session.add(new_user)
        db.session.commit()
        db.session.refresh(new_user, ["champs_autorises"])
        return {
            "id": new_user.id,
            "username": new_user.username,
            "is_admin": new_user.is_admin,
            "is_dashboard_only": new_user.is_dashboard_only,
            "allowed_champs": [c.champno for c in new_user.champs_autorises],
        }
    except IntegrityError:
        db.session.rollback()
        raise DuplicateEntityError("Ce nom d'utilisateur est déjà pris.")
    except Exception as e:
        db.session.rollback()
        raise ServiceException(f"Erreur ORM lors de la création de l'utilisateur: {e}")


def update_user_role_service(user_id: int, role: str, allowed_champs: list[str]) -> None:
    """Met à jour le rôle et les accès d'un utilisateur en utilisant l'ORM."""
    user = db.session.get(User, user_id)
    if not user:
        raise EntityNotFoundError("Utilisateur non trouvé.")

    user.is_admin = False
    user.is_dashboard_only = False
    user.champs_autorises.clear()

    if role == "admin":
        user.is_admin = True
    elif role == "dashboard_only":
        user.is_dashboard_only = True
    elif role == "specific_champs" and allowed_champs:
        champs_objects = db.session.query(Champ).filter(Champ.champno.in_(allowed_champs)).all()
        if len(champs_objects) != len(allowed_champs):
            raise BusinessRuleValidationError("Un ou plusieurs champs spécifiés sont invalides.")
        user.champs_autorises = champs_objects

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        raise ServiceException(f"La mise à jour du rôle a échoué: {e}")


def delete_user_service(user_id_to_delete: int, current_user_id: int) -> None:
    """Supprime un utilisateur en utilisant l'ORM, avec des vérifications de règles métier."""
    if user_id_to_delete == current_user_id:
        raise BusinessRuleValidationError("Vous ne pouvez pas vous supprimer vous-même.")

    user_to_delete = db.session.get(User, user_id_to_delete)
    if not user_to_delete:
        raise EntityNotFoundError("Utilisateur non trouvé.")

    if user_to_delete.is_admin:
        admin_count = db.session.query(User).filter_by(is_admin=True).count()
        if admin_count <= 1:
            raise BusinessRuleValidationError("Impossible de supprimer le dernier administrateur.")

    try:
        db.session.delete(user_to_delete)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        raise ServiceException(f"La suppression de l'utilisateur a échoué: {e}")


# --- Services pour les Attributions ---


def get_attribution_details_service(attribution_id: int) -> dict[str, Any]:
    """Récupère les détails d'une attribution, y compris le champno de l'enseignant."""
    attribution = db.session.get(AttributionCours, attribution_id)
    if not attribution:
        raise EntityNotFoundError("Attribution non trouvée.")

    return {
        "attributionid": attribution.attributionid,
        "enseignantid": attribution.enseignantid,
        "champno": attribution.enseignant.champno,
        "estfictif": attribution.enseignant.estfictif,
    }


def get_remaining_groups_for_course_service(code_cours: str, annee_id: int) -> int:
    """
    Calcule le nombre de groupes restants pour un cours via l'ORM.
    Utilisé par l'API pour mettre à jour l'interface utilisateur.
    """
    cours = db.session.get(Cours, {"codecours": code_cours, "annee_id": annee_id})
    if not cours:
        # Retourne 0 si le cours n'existe pas, car il n'y a pas de groupes.
        return 0

    groupes_pris = db.session.query(func.coalesce(func.sum(AttributionCours.nbgroupespris), 0)).filter_by(codecours=code_cours, annee_id_cours=annee_id).scalar()

    return cours.nbgroupeinitial - (groupes_pris or 0)


def add_attribution_service(enseignant_id: int, code_cours: str, annee_id: int) -> int:
    """Ajoute une attribution de cours à un enseignant via l'ORM."""
    # 1. Valider les entités et les règles métier
    enseignant = db.session.get(Enseignant, enseignant_id)
    if not enseignant:
        raise EntityNotFoundError("Enseignant non trouvé.")

    cours = db.session.get(Cours, {"codecours": code_cours, "annee_id": annee_id})
    if not cours:
        raise EntityNotFoundError("Cours non trouvé.")

    # Règle: Vérifier le verrouillage du champ
    statut_champ = db.session.query(ChampAnneeStatut).filter_by(champ_no=enseignant.champno, annee_id=annee_id).first()
    if statut_champ and statut_champ.est_verrouille and not enseignant.estfictif:
        raise BusinessRuleValidationError("Les modifications sont désactivées car le champ est verrouillé.")

    # Règle: Vérifier les groupes restants
    groupes_restants = get_remaining_groups_for_course_service(code_cours, annee_id)
    if groupes_restants < 1:
        raise BusinessRuleValidationError("Plus de groupes disponibles pour ce cours.")

    # 2. Exécuter l'opération
    try:
        new_attribution = AttributionCours(
            enseignantid=enseignant_id,
            codecours=code_cours,
            annee_id_cours=annee_id,
            nbgroupespris=1,
        )
        db.session.add(new_attribution)
        db.session.commit()
        return new_attribution.attributionid
    except Exception as e:
        db.session.rollback()
        raise ServiceException(f"Erreur de base de données lors de l'attribution : {e}")


def delete_attribution_service(attribution_id: int) -> dict[str, Any]:
    """Supprime une attribution de cours via l'ORM."""
    # 1. Récupérer l'entité et valider les règles métier
    attribution = (
        db.session.query(AttributionCours)
        .options(joinedload(AttributionCours.enseignant).joinedload(Enseignant.champ).joinedload(Champ.statuts_annee))
        .filter(AttributionCours.attributionid == attribution_id)
        .first()
    )

    if not attribution:
        raise EntityNotFoundError("Attribution non trouvée.")

    enseignant = attribution.enseignant
    annee_id = enseignant.annee_id

    # Règle: Vérifier le verrouillage du champ
    statut_champ = db.session.query(ChampAnneeStatut).filter_by(champ_no=enseignant.champno, annee_id=annee_id).first()

    if statut_champ and statut_champ.est_verrouille and not enseignant.estfictif:
        raise BusinessRuleValidationError("Les modifications sont désactivées car le champ est verrouillé.")

    # 2. Préparer les données à retourner (contrat de la fonction)
    attr_info = {
        "enseignantid": enseignant.enseignantid,
        "codecours": attribution.codecours,
        "annee_id_cours": attribution.annee_id_cours,
    }

    # 3. Exécuter l'opération
    try:
        db.session.delete(attribution)
        db.session.commit()
        return attr_info
    except Exception as e:
        db.session.rollback()
        raise ServiceException(f"Échec de la suppression de l'attribution : {e}")


def _get_all_cours_grouped_by_champ_orm(annee_id: int) -> dict[str, dict[str, Any]]:
    """Récupère tous les cours d'une année via l'ORM, regroupés par champ."""
    cours_list = db.session.query(Cours).options(joinedload(Cours.champ)).filter(Cours.annee_id == annee_id).order_by(Cours.champno, Cours.codecours).all()

    cours_par_champ: defaultdict[str, dict[str, Any]] = defaultdict(lambda: {"champ_nom": "", "cours": []})
    for cours in cours_list:
        champ_no = cours.champno
        if not cours_par_champ[champ_no]["champ_nom"]:
            cours_par_champ[champ_no]["champ_nom"] = cours.champ.champnom
        cours_par_champ[champ_no]["cours"].append(_cours_to_dict(cours))
    return dict(cours_par_champ)


def _get_all_enseignants_grouped_by_champ_orm(annee_id: int) -> dict[str, dict[str, Any]]:
    """Récupère tous les enseignants d'une année via l'ORM, regroupés par champ."""
    enseignants_list = (
        db.session.query(Enseignant)
        .options(joinedload(Enseignant.champ))
        .filter(Enseignant.annee_id == annee_id, Enseignant.estfictif == False)  # noqa: E712
        .order_by(Enseignant.champno, Enseignant.nom, Enseignant.prenom)
        .all()
    )

    enseignants_par_champ: defaultdict[str, dict[str, Any]] = defaultdict(lambda: {"champ_nom": "", "enseignants": []})
    for enseignant in enseignants_list:
        champ_no = enseignant.champno
        if not enseignants_par_champ[champ_no]["champ_nom"]:
            enseignants_par_champ[champ_no]["champ_nom"] = enseignant.champ.champnom
        enseignants_par_champ[champ_no]["enseignants"].append(_enseignant_to_dict(enseignant))
    return dict(enseignants_par_champ)


def _create_teacher_sort_key(teacher_data: dict[str, Any]) -> tuple:
    """Crée une clé de tri complexe pour les enseignants et les tâches."""
    is_fictif = teacher_data["estfictif"]
    nom_complet = teacher_data["nomcomplet"]

    sort_order = 0
    if is_fictif:
        if "Tâche restante" in nom_complet:
            sort_order = 1
        elif "Non attribué" in nom_complet:
            sort_order = 2
        else:
            sort_order = 3
    task_num = 0
    if sort_order == 1:
        try:
            task_num = int(nom_complet.split("-")[-1])
        except (ValueError, IndexError):
            pass
    nom = teacher_data.get("nom", "") or ""
    prenom = teacher_data.get("prenom", "") or ""
    return (sort_order, task_num, nom, prenom)


# --- Services d'agrégation de données pour les vues ---


def get_data_for_admin_page_service(annee_id: int) -> dict[str, Any]:
    """Récupère toutes les données nécessaires pour la page d'administration via l'ORM."""
    try:
        return {
            "cours_par_champ": _get_all_cours_grouped_by_champ_orm(annee_id),
            "enseignants_par_champ": _get_all_enseignants_grouped_by_champ_orm(annee_id),
            "tous_les_champs": get_all_champs_service(),
            "tous_les_financements": get_all_financements_service(),
        }
    except Exception as e:
        raise ServiceException(f"Erreur lors de l'agrégation des données pour la page admin: {e}")


def get_data_for_user_admin_page_service() -> dict[str, Any]:
    """Récupère toutes les données nécessaires pour la page d'administration des utilisateurs."""
    try:
        user_details = get_all_users_with_details_service()
        all_champs = get_all_champs_service()

        return {
            "users": user_details["users"],
            "all_champs": all_champs,
        }
    except Exception as e:
        raise ServiceException(f"Erreur lors de l'agrégation des données utilisateur : {e}")


def get_data_for_champ_page_service(champ_no: str, annee_id: int) -> dict[str, Any]:
    """
    Récupère et agrège toutes les données nécessaires pour la page de détail d'un champ.
    """
    try:
        # 1. Récupérer les détails du champ (lève une exception si non trouvé)
        champ_details = get_champ_details_service(champ_no, annee_id)

        # 2. Récupérer tous les enseignants de l'année (requête optimisée unique)
        all_teachers_details = _get_all_teachers_with_details_service(annee_id)
        # Filtrer en Python pour ne garder que ceux du champ concerné
        enseignants_du_champ = [teacher for teacher in all_teachers_details if teacher["champno"] == champ_no]
        sorted_enseignants = sorted(enseignants_du_champ, key=_create_teacher_sort_key)

        # 3. Récupérer tous les cours du champ pour l'année
        cours_du_champ_orm = db.session.query(Cours).filter_by(champno=champ_no, annee_id=annee_id).order_by(Cours.codecours).all()
        cours_du_champ = [_cours_to_dict(c) for c in cours_du_champ_orm]

        # 4. Récupérer tous les champs pour les menus déroulants
        tous_les_champs = get_all_champs_service()

        return {
            "champ": champ_details,
            "enseignants": sorted_enseignants,
            "cours": cours_du_champ,
            "tous_les_champs": tous_les_champs,
        }
    except EntityNotFoundError as e:
        # Laisser passer cette exception pour que la vue puisse la gérer (ex: 404)
        raise e
    except Exception as e:
        # Encapsuler les autres erreurs
        raise ServiceException(f"Erreur lors de l'agrégation des données pour la page du champ {champ_no}: {e}")


def get_dashboard_summary_service(annee_id: int) -> dict[str, Any]:
    """Service principal qui orchestre la récupération des données du tableau de bord."""
    try:
        return get_dashboard_summary_service_orm(annee_id)
    except Exception as e:
        raise ServiceException(f"Erreur lors du calcul du sommaire du tableau de bord: {e}")


def get_dashboard_summary_service_orm(annee_id: int) -> dict[str, Any]:
    """Calcule et récupère les données du tableau de bord en utilisant SQLAlchemy ORM."""
    ac_alias = aliased(AttributionCours)
    c_alias = aliased(Cours)
    teacher_periods_subquery = (
        db.session.query(Enseignant.enseignantid, func.coalesce(func.sum(c_alias.nbperiodes * ac_alias.nbgroupespris), 0).label("total_periodes"))
        .outerjoin(ac_alias, Enseignant.enseignantid == ac_alias.enseignantid)
        .outerjoin(c_alias, (ac_alias.codecours == c_alias.codecours) & (ac_alias.annee_id_cours == c_alias.annee_id))
        .filter(Enseignant.annee_id == annee_id)
        .group_by(Enseignant.enseignantid)
        .subquery()
    )
    tps_alias = aliased(teacher_periods_subquery, name="tps")

    is_tp_non_fictif = (Enseignant.esttempsplein == True) & (Enseignant.estfictif == False)  # noqa: E712
    champ_stats_query = (
        db.session.query(
            Enseignant.champno,
            func.count(case((is_tp_non_fictif, Enseignant.enseignantid), else_=None)).label("nb_enseignants_tp"),
            func.coalesce(func.sum(case((is_tp_non_fictif, tps_alias.c.total_periodes), else_=0)), 0).label("periodes_choisies_tp"),
        )
        .join(tps_alias, Enseignant.enseignantid == tps_alias.c.enseignantid)
        .filter(Enseignant.annee_id == annee_id)
        .group_by(Enseignant.champno)
        .subquery()
    )
    cs_alias = aliased(champ_stats_query, name="cs")

    all_champs_data = (
        db.session.query(
            Champ.champno,
            Champ.champnom,
            func.coalesce(ChampAnneeStatut.est_verrouille, False).label("est_verrouille"),
            func.coalesce(ChampAnneeStatut.est_confirme, False).label("est_confirme"),
            func.coalesce(cs_alias.c.nb_enseignants_tp, 0).label("nb_enseignants_tp"),
            func.coalesce(cs_alias.c.periodes_choisies_tp, 0).label("periodes_choisies_tp"),
        )
        .outerjoin(cs_alias, Champ.champno == cs_alias.c.champno)
        .outerjoin(ChampAnneeStatut, (Champ.champno == ChampAnneeStatut.champ_no) & (ChampAnneeStatut.annee_id == annee_id))
        .order_by(Champ.champno)
        .all()
    )

    moyennes_par_champ = {}
    for row in all_champs_data:
        nb_tp = int(row.nb_enseignants_tp)
        periodes_tp = float(row.periodes_choisies_tp)
        moyenne = (periodes_tp / nb_tp) if nb_tp > 0 else 0.0
        moyennes_par_champ[row.champno] = {
            "champ_nom": row.champnom,
            "est_verrouille": row.est_verrouille,
            "est_confirme": row.est_confirme,
            "nb_enseignants_tp": nb_tp,
            "periodes_choisies_tp": periodes_tp,
            "moyenne": moyenne,
            "periodes_magiques": periodes_tp - (nb_tp * 24),
        }

    total_periodes_global_tp = sum(r["periodes_choisies_tp"] for r in moyennes_par_champ.values())
    nb_enseignants_tp_global = sum(r["nb_enseignants_tp"] for r in moyennes_par_champ.values())
    total_periodes_confirme_tp = sum(r["periodes_choisies_tp"] for r in moyennes_par_champ.values() if r["est_confirme"])
    nb_enseignants_confirme_tp = sum(r["nb_enseignants_tp"] for r in moyennes_par_champ.values() if r["est_confirme"])
    moyenne_generale = (total_periodes_global_tp / nb_enseignants_tp_global) if nb_enseignants_tp_global > 0 else 0.0
    moyenne_prelim_conf = (total_periodes_confirme_tp / nb_enseignants_confirme_tp) if nb_enseignants_confirme_tp > 0 else 0.0
    grand_totals = {
        "total_enseignants_tp": float(nb_enseignants_tp_global),
        "total_periodes_choisies_tp": total_periodes_global_tp,
        "total_periodes_magiques": sum(r["periodes_magiques"] for r in moyennes_par_champ.values()),
    }
    return {
        "moyennes_par_champ": moyennes_par_champ,
        "moyenne_generale": moyenne_generale,
        "moyenne_preliminaire_confirmee": moyenne_prelim_conf,
        "grand_totals": grand_totals,
    }


def _calculate_teacher_details(enseignant: Enseignant) -> dict[str, Any]:
    """
    Fonction helper pour calculer les détails (périodes, attributions) pour UN enseignant.
    L'enseignant doit être pré-chargé avec ses attributions et cours via joinedload.
    """
    periodes_cours = 0.0
    periodes_autres = 0.0
    attributions_details = []

    for attr in enseignant.attributions:
        # La relation `attr.cours` devrait déjà être chargée
        cours_info = attr.cours
        periodes_a_ajouter = float(cours_info.nbperiodes) * attr.nbgroupespris

        if cours_info.estcoursautre:
            periodes_autres += periodes_a_ajouter
        else:
            periodes_cours += periodes_a_ajouter

        attributions_details.append(
            {
                "AttributionID": attr.attributionid,
                "CodeCours": cours_info.codecours,
                "NbGroupesPris": attr.nbgroupespris,
                "CoursDescriptif": cours_info.coursdescriptif,
                "NbPeriodes": float(cours_info.nbperiodes),
                "EstCoursAutre": cours_info.estcoursautre,
                "annee_id": cours_info.annee_id,
                "financement_code": cours_info.financement_code,
            }
        )
    return {
        "attributions": attributions_details,
        "periodes": {
            "periodes_cours": periodes_cours,
            "periodes_autres": periodes_autres,
            "total_periodes": periodes_cours + periodes_autres,
        },
    }


def _get_all_teachers_with_details_service(annee_id: int) -> list[dict[str, Any]]:
    """
    Récupère tous les enseignants d'une année avec leurs attributions et calculs de périodes.
    Utilise l'ORM avec chargement optimisé pour éviter les requêtes N+1.
    """
    try:
        enseignants = (
            db.session.query(Enseignant)
            .options(joinedload(Enseignant.attributions).joinedload(AttributionCours.cours))
            .filter(Enseignant.annee_id == annee_id)
            .order_by(Enseignant.estfictif, Enseignant.nom, Enseignant.prenom)
            .all()
        )
        results = []
        for enseignant in enseignants:
            details = _calculate_teacher_details(enseignant)
            results.append({**_enseignant_to_dict(enseignant), **details})
        return results
    except Exception as e:
        raise ServiceException(f"Erreur ORM lors de la récupération des détails des enseignants : {e}")


def get_teacher_update_payload_service(enseignant_id: int) -> dict[str, Any]:
    """
    Récupère les données nécessaires (périodes, attributions) pour mettre à jour
    l'interface d'un enseignant après une action.
    """
    try:
        enseignant = db.session.query(Enseignant).options(joinedload(Enseignant.attributions).joinedload(AttributionCours.cours)).filter(Enseignant.enseignantid == enseignant_id).one_or_none()

        if not enseignant:
            raise EntityNotFoundError("Enseignant non trouvé.")

        details = _calculate_teacher_details(enseignant)
        # Le format de réponse attendu par l'API est directement le contenu du dictionnaire de détails
        return {
            "periodes_enseignant": details["periodes"],
            "attributions_enseignant": details["attributions"],
        }
    except Exception as e:
        raise ServiceException(f"Erreur ORM lors de la récupération du payload de mise à jour de l'enseignant : {e}")


def get_detailed_tasks_data_service(annee_id: int) -> list[dict[str, Any]]:
    """
    Orchestre la récupération et le formatage des données pour la page principale des tâches.
    Utilise maintenant entièrement l'ORM.
    """
    tous_les_enseignants_details = _get_all_teachers_with_details_service(annee_id)
    tous_les_champs = get_all_champs_service()
    statuts_champs = get_all_champ_statuses_for_year_service(annee_id)
    enseignants_par_champ_temp: dict[str, dict[str, Any]] = {
        str(champ["champno"]): {
            "champno": str(champ["champno"]),
            "champnom": champ["champnom"],
            "enseignants": [],
            "est_verrouille": statuts_champs.get(str(champ["champno"]), {}).get("est_verrouille", False),
            "est_confirme": statuts_champs.get(str(champ["champno"]), {}).get("est_confirme", False),
        }
        for champ in tous_les_champs
    }
    for ens in tous_les_enseignants_details:
        champ_no = ens["champno"]
        if champ_no in enseignants_par_champ_temp:
            enseignants_par_champ_temp[champ_no]["enseignants"].append(ens)
    return list(enseignants_par_champ_temp.values())


def get_attributions_for_export_service(annee_id: int) -> dict[str, dict[str, Any]]:
    """
    Récupère les attributions formatées pour l'export via l'ORM et les groupe par champ.
    Cette version est entièrement refactorisée pour utiliser SQLAlchemy ORM.
    """
    try:
        # Étape 1 : Requête ORM pour récupérer toutes les données nécessaires en une seule fois.
        query_results = (
            db.session.query(
                Champ.champno.label("champno"),
                Champ.champnom.label("champnom"),
                Enseignant.nom.label("nom"),
                Enseignant.prenom.label("prenom"),
                Cours.codecours.label("codecours"),
                Cours.coursdescriptif.label("coursdescriptif"),
                Cours.estcoursautre.label("estcoursautre"),
                Cours.financement_code.label("financement_code"),
                func.sum(AttributionCours.nbgroupespris).label("total_groupes_pris"),
                Cours.nbperiodes.label("nbperiodes"),
            )
            .select_from(AttributionCours)
            .join(AttributionCours.enseignant)
            .join(AttributionCours.cours)
            .join(Cours.champ)
            .filter(Enseignant.annee_id == annee_id, Enseignant.estfictif == False)  # noqa: E712
            .group_by(
                Champ.champno,
                Champ.champnom,
                Enseignant.nom,
                Enseignant.prenom,
                Cours.codecours,
                Cours.coursdescriptif,
                Cours.estcoursautre,
                Cours.financement_code,
                Cours.nbperiodes,
            )
            .order_by(Champ.champno, Enseignant.nom, Enseignant.prenom, Cours.codecours)
            .all()
        )
        # Étape 2 : Conversion des résultats en une liste de dictionnaires pour un traitement facile.
        attributions_raw = [dict(row._mapping) for row in query_results]

    except Exception as e:
        raise ServiceException(f"Erreur ORM lors de la récupération des attributions pour l'export: {e}")

    if not attributions_raw:
        return {}

    # Étape 3 : Formatage des données dans la structure attendue, regroupée par champ.
    attributions_par_champ: defaultdict[str, dict[str, Any]] = defaultdict(lambda: {"nom": "", "attributions": []})
    for attr in attributions_raw:
        champ_no = attr["champno"]
        if not attributions_par_champ[champ_no]["nom"]:
            attributions_par_champ[champ_no]["nom"] = attr["champnom"]

        attr_copy = attr.copy()
        # Assurer que le type de données est correct pour les calculs futurs.
        attr_copy["nbperiodes"] = float(attr_copy["nbperiodes"])
        attributions_par_champ[champ_no]["attributions"].append(attr_copy)

    return dict(attributions_par_champ)


def get_periodes_restantes_for_export_service(annee_id: int) -> dict[str, dict[str, Any]]:
    """
    Récupère les périodes restantes (attributions à des enseignants fictifs) pour l'export.
    Cette version est entièrement refactorisée pour utiliser SQLAlchemy ORM.
    """
    try:
        # Requête ORM pour récupérer les attributions des enseignants fictifs.
        query_results = (
            db.session.query(
                Enseignant.champno.label("champno"),
                Champ.champnom.label("champnom"),
                Enseignant.nomcomplet.label("nomcomplet"),
                Cours.codecours.label("codecours"),
                Cours.coursdescriptif.label("coursdescriptif"),
                func.sum(AttributionCours.nbgroupespris).label("total_groupes_pris"),
                Cours.nbperiodes.label("nbperiodes"),
            )
            .select_from(AttributionCours)
            .join(AttributionCours.enseignant)
            .join(Enseignant.champ)
            .join(AttributionCours.cours)
            .filter(Enseignant.annee_id == annee_id, Enseignant.estfictif == True)  # noqa: E712
            .group_by(
                Enseignant.champno,
                Champ.champnom,
                Enseignant.nomcomplet,
                Cours.codecours,
                Cours.coursdescriptif,
                Cours.nbperiodes,
            )
            .order_by(Enseignant.champno.asc(), Enseignant.nomcomplet.asc(), Cours.codecours.asc())
            .all()
        )
        periodes_restantes_raw = [dict(row._mapping) for row in query_results]
    except Exception as e:
        raise ServiceException(f"Erreur ORM lors de la récupération des périodes restantes pour l'export: {e}")

    if not periodes_restantes_raw:
        return {}

    # Formatage des données dans la structure attendue, regroupée par champ.
    periodes_par_champ: defaultdict[str, dict[str, Any]] = defaultdict(lambda: {"nom": "", "periodes": []})
    for periode in periodes_restantes_raw:
        champ_no = periode["champno"]
        if not periodes_par_champ[champ_no]["nom"]:
            periodes_par_champ[champ_no]["nom"] = periode["champnom"]

        periode_copy = periode.copy()
        periode_copy["nbperiodes"] = float(periode_copy["nbperiodes"])
        periodes_par_champ[champ_no]["periodes"].append(periode_copy)

    return dict(periodes_par_champ)


def get_org_scolaire_export_data_service(annee_id: int) -> dict[str, dict[str, Any]]:
    """
    Construit les données pour l'export "Organisation Scolaire" via l'ORM et Python.
    """
    try:
        all_financements = db.session.query(TypeFinancement).all()
        header_map = {f"PÉRIODES {f.libelle.upper()}": f.code for f in all_financements}
        code_to_header_map = {code: header for header, code in header_map.items()}
        all_headers = sorted(list(header_map.keys()))
        all_headers.insert(0, "PÉRIODES RÉGULIER")

        pivot_data: defaultdict[str, dict[str, Any]] = defaultdict(lambda: {})
        all_teachers = db.session.query(Enseignant).options(joinedload(Enseignant.champ)).filter_by(annee_id=annee_id).all()

        for enseignant in all_teachers:
            champ = enseignant.champ
            enseignant_key = f"{'fictif' if enseignant.estfictif else 'reel'}-{enseignant.nomcomplet}"
            display_nomcomplet = enseignant.nomcomplet
            if enseignant.estfictif and enseignant.nomcomplet.startswith(f"{champ.champno}-"):
                display_nomcomplet = display_nomcomplet.replace(f"{champ.champno}-", "", 1).strip()
            pivot_data[champ.champno][enseignant_key] = {
                "nom": enseignant.nom,
                "prenom": enseignant.prenom,
                "nomcomplet": display_nomcomplet,
                "estfictif": enseignant.estfictif,
                "champnom": champ.champnom,
                **{header: 0.0 for header in all_headers},
            }

        attributions = (
            db.session.query(AttributionCours).options(joinedload(AttributionCours.enseignant), joinedload(AttributionCours.cours)).join(Enseignant).filter(Enseignant.annee_id == annee_id).all()
        )
        for attr in attributions:
            enseignant = attr.enseignant
            cours = attr.cours
            enseignant_key = f"{'fictif' if enseignant.estfictif else 'reel'}-{enseignant.nomcomplet}"
            if enseignant_key in pivot_data[enseignant.champno]:
                total_p = float(cours.nbperiodes) * attr.nbgroupespris
                target_col = code_to_header_map.get(cours.financement_code, "PÉRIODES RÉGULIER")
                pivot_data[enseignant.champno][enseignant_key][target_col] += total_p

        cours_with_groups_taken = (
            db.session.query(AttributionCours.codecours, func.sum(AttributionCours.nbgroupespris).label("groupes_pris"))
            .filter(AttributionCours.annee_id_cours == annee_id)
            .group_by(AttributionCours.codecours)
            .all()
        )
        groups_taken_map = {c.codecours: c.groupes_pris for c in cours_with_groups_taken}
        all_courses = db.session.query(Cours).options(joinedload(Cours.champ)).filter_by(annee_id=annee_id).all()

        unassigned_tasks: defaultdict[str, dict[str, Any]] = defaultdict(lambda: defaultdict(float))
        for cours in all_courses:
            remaining_groups = cours.nbgroupeinitial - groups_taken_map.get(cours.codecours, 0)
            if remaining_groups > 0:
                total_p = float(cours.nbperiodes) * remaining_groups
                target_col = code_to_header_map.get(cours.financement_code, "PÉRIODES RÉGULIER")
                unassigned_tasks[cours.champno][target_col] += total_p

        # CORRIGÉ : Logique revue pour ajouter la ligne "Non attribué" correctement
        for champ_no, periods_by_funding in unassigned_tasks.items():
            if any(p > 0 for p in periods_by_funding.values()):
                unassigned_key = "fictif-Non attribué"
                champ = db.session.get(Champ, champ_no)

                # Initialise le dictionnaire du champ s'il n'existe pas encore
                if champ_no not in pivot_data:
                    pivot_data[champ_no] = {}

                # Initialise ou met à jour la ligne "Non attribué"
                if unassigned_key not in pivot_data[champ_no]:
                    pivot_data[champ_no][unassigned_key] = {
                        "nom": "Non attribué",
                        "prenom": None,
                        "nomcomplet": "Non attribué",
                        "estfictif": True,
                        "champnom": champ.champnom if champ else "N/A",
                        **{header: 0.0 for header in all_headers},
                    }

                for header, value in periods_by_funding.items():
                    pivot_data[champ_no][unassigned_key][header] += value

        donnees_par_champ: dict[str, dict[str, Any]] = {}
        for champ_no, enseignants_data in pivot_data.items():
            filtered_enseignants = [data for data in enseignants_data.values() if sum(data.get(h, 0.0) for h in all_headers) > 0 or not data["estfictif"]]
            if not filtered_enseignants:
                continue
            sorted_enseignants = sorted(filtered_enseignants, key=_create_teacher_sort_key)
            donnees_par_champ[champ_no] = {"nom": sorted_enseignants[0]["champnom"], "donnees": sorted_enseignants}
        return dict(sorted(donnees_par_champ.items()))

    except Exception as e:
        raise ServiceException(f"Erreur ORM lors de la génération de l'export Organisation Scolaire: {e}")


def get_preparation_horaire_data_service(annee_id: int) -> dict[str, Any]:
    """
    Récupère les données pour la préparation de l'horaire via l'ORM SQLAlchemy.
    """
    try:
        all_champs = get_all_champs_service()
        all_cours_raw = db.session.query(Cours).filter_by(annee_id=annee_id).order_by(Cours.champno, Cours.codecours).all()
        all_assignments_raw = (
            db.session.query(AttributionCours)
            .join(AttributionCours.enseignant)
            .filter(Enseignant.annee_id == annee_id, Enseignant.estfictif == False)  # noqa: E712
            .options(joinedload(AttributionCours.enseignant), joinedload(AttributionCours.cours))
            .order_by(AttributionCours.codecours, Enseignant.nomcomplet)
            .all()
        )
        saved_assignments_raw = db.session.query(PreparationHoraire).filter_by(annee_id=annee_id).all()

        cours_par_champ: defaultdict[str, list] = defaultdict(list)
        cours_details: dict[str, dict] = {}
        for cours in all_cours_raw:
            cours_dict = {"codecours": cours.codecours, "champno": cours.champno, "annee_id": cours.annee_id}
            cours_par_champ[cours.champno].append(cours_dict)
            cours_details[cours.codecours] = cours_dict

        enseignants_par_cours: defaultdict[str, list] = defaultdict(list)
        for assignment in all_assignments_raw:
            enseignant = assignment.enseignant
            for _ in range(assignment.nbgroupespris):
                enseignants_par_cours[assignment.codecours].append({"codecours": assignment.codecours, "nomcomplet": enseignant.nomcomplet, "enseignantid": enseignant.enseignantid})

        saved_placements = defaultdict(lambda: defaultdict(list))
        for saved in saved_assignments_raw:
            key = (saved.secondaire_level, saved.codecours)
            saved_placements[key][saved.colonne_assignee].append(saved.enseignant_id)

        prepared_grid: dict[int, list] = {level: [] for level in range(1, 6)}
        for (level, codecours), columns in saved_placements.items():
            cours_info = cours_details.get(codecours)
            if not cours_info:
                continue
            all_teachers_for_course = enseignants_par_cours.get(codecours, [])
            teachers_lookup = {t["enseignantid"]: t for t in all_teachers_for_course}
            unassigned_teachers = []
            if all_teachers_for_course:
                placed_assignments_count = sum(len(ids) for ids in columns.values())
                unassigned_count = len(all_teachers_for_course) - placed_assignments_count
                if unassigned_count > 0:
                    teacher_prototype = all_teachers_for_course[0]
                    unassigned_teachers = [teacher_prototype] * unassigned_count
            prepared_grid[level].append(
                {
                    "cours": cours_info,
                    "all_teachers_for_course": all_teachers_for_course,
                    "unassigned_teachers": unassigned_teachers,
                    "assigned_teachers_by_col": columns,
                    "teachers_lookup": teachers_lookup,
                }
            )
        return {
            "all_champs": all_champs,
            "cours_par_champ": dict(cours_par_champ),
            "enseignants_par_cours": dict(enseignants_par_cours),
            "prepared_grid": prepared_grid,
        }
    except Exception as e:
        raise ServiceException(f"Erreur ORM lors de la préparation des données pour l'horaire : {e}")


def save_preparation_horaire_service(annee_id: int, assignments_data: list[dict[str, Any]]) -> None:
    """
    Sauvegarde (en remplaçant) les assignations de la préparation de l'horaire pour une année.
    """
    required_keys = ["secondaire_level", "codecours", "annee_id_cours", "enseignant_id", "colonne_assignee"]
    for item in assignments_data:
        if not all(key in item for key in required_keys):
            raise BusinessRuleValidationError("Données de sauvegarde invalides ou incomplètes.")

    try:
        db.session.query(PreparationHoraire).filter_by(annee_id=annee_id).delete(synchronize_session=False)
        if assignments_data:
            new_assignments = [PreparationHoraire(annee_id=annee_id, **data) for data in assignments_data]
            db.session.add_all(new_assignments)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        raise ServiceException(f"Erreur ORM lors de la sauvegarde de la préparation de l'horaire : {e}")
